"""
Модуль для работы с Excel файлами.
Поддерживает чтение существующих файлов, перезапись листов, запись данных
и работу с шаблонами расчетов.
"""

import os
import re
import copy
from typing import List, Optional, Dict, Tuple
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ExcelHandler:
    """Обработчик Excel файлов."""
    
    def __init__(self, file_path: str):
        """
        Инициализирует обработчик Excel файла.
        
        Args:
            file_path: Путь к Excel файлу
        """
        self.file_path = file_path
        self.workbook = None
    
    def load_workbook(self) -> None:
        """Загружает существующий Excel файл или создает новый."""
        if os.path.exists(self.file_path):
            self.workbook = openpyxl.load_workbook(self.file_path)
        else:
            self.workbook = openpyxl.Workbook()
    
    def get_sheet_names(self) -> List[str]:
        """Возвращает список имен листов в книге."""
        if self.workbook is None:
            self.load_workbook()
        return self.workbook.sheetnames
    
    def overwrite_sheet(
        self,
        sheet_name: str,
        headers: List[str],
        rows: List[List[str]],
        start_row: int = 1
    ) -> None:
        """
        Перезаписывает указанный лист в Excel файле.
        Если лист не существует, создает новый.
        
        Args:
            sheet_name: Имя листа
            headers: Заголовки колонок
            rows: Строки данных
            start_row: Начальная строка (по умолчанию 1)
        """
        if self.workbook is None:
            self.load_workbook()
        
        # Удаляем существующий лист с таким именем
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
        
        # Создаем новый лист
        worksheet = self.workbook.create_sheet(title=sheet_name)
        
        # Записываем заголовки
        if headers:
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=start_row, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="FFFFFF")
        
        # Записываем строки данных
        for row_idx, row_data in enumerate(rows, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')
        
        # Автоматическая ширина колонок
        self._auto_adjust_column_width(worksheet)
    
    def save(self, output_path: Optional[str] = None) -> None:
        """
        Сохраняет Excel файл.
        
        Args:
            output_path: Путь для сохранения (если None, сохраняет в исходный файл)
        """
        if self.workbook is None:
            raise ValueError("Книга не загружена. Вызовите load_workbook() сначала.")
        
        # Устанавливаем флаг автоматического пересчета формул
        # Это заставит Excel пересчитать все формулы при открытии файла
        self.workbook.calculation.calcOnSave = True
        self.workbook.calculation.calcId = 0  # Сбрасываем ID расчета
        self.workbook.calculation.fullCalcOnLoad = True  # Полный пересчет при загрузке
        
        save_path = output_path or self.file_path
        self.workbook.save(save_path)
    
    def _auto_adjust_column_width(self, worksheet) -> None:
        """Автоматически настраивает ширину колонок."""
        for column_cells in worksheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)
            
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass
            
            # Добавляем небольшой отступ
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            worksheet.column_dimensions[column].width = max(adjusted_width, 10)  # Минимум 10 символов

    def write_data_columns_pu(
        self,
        sheet_name: str,
        data_rows: List[List],
        start_row: int = 11
    ) -> int:
        """
        Записывает данные в колонки P-U (16-21) для использования в формулах графиков.
        
        Колонки:
          P (16) - Pвых дБм
          Q (17) - Ig мА
          R (18) - Id А
          S (19) - Gain dB
          T (20) - КПД %
          U (21) - Pвых Вт
        
        Args:
            sheet_name: Имя листа
            data_rows: Строки данных из TXT
            start_row: Начальная строка (по умолчанию 11)
            
        Returns:
            Количество записанных строк
        """
        if self.workbook is None:
            self.load_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
        
        ws = self.workbook[sheet_name]
        
        rows_written = 0
        for i, row_data in enumerate(data_rows):
            target_row = start_row + i
            # Порядок в row_data:
            # [0]=№, [1]=f MHz, [2]=P dBm, [3]=IG mA, [4]=ID A, [5]=Gain dB, [6]=КПД %, [7]=Pвых W
            if len(row_data) >= 8:
                ws.cell(row=target_row, column=16, value=self._convert_value(row_data[2]))  # P - PdBm
                ws.cell(row=target_row, column=17, value=self._convert_value(row_data[3]))  # Q - Ig mA
                ws.cell(row=target_row, column=18, value=self._convert_value(row_data[4]))  # R - Id A
                ws.cell(row=target_row, column=19, value=self._convert_value(row_data[5]))  # S - Gain dB
                ws.cell(row=target_row, column=20, value=self._convert_value(row_data[6]))  # T - Efficiency %
                ws.cell(row=target_row, column=21, value=self._convert_value(row_data[7]))  # U - Pout W
            rows_written += 1

        first_empty_row = start_row + rows_written
        if first_empty_row <= ws.max_row:
            self._clear_columns_range(ws, first_empty_row, ws.max_row, range(16, 22))
        
        return rows_written

    def copy_template_sheet(
        self,
        source_sheet_name: str,
        new_sheet_name: str
    ) -> None:
        """
        Копирует шаблон листа с сохранением формул и структуры.
        Использует ручное копирование для корректного сохранения формул.
        
        Args:
            source_sheet_name: Имя исходного листа-шаблона (например, '№1')
            new_sheet_name: Имя нового листа
        """
        if self.workbook is None:
            self.load_workbook()
        
        if source_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Исходный лист '{source_sheet_name}' не найден")
        
        source_ws = self.workbook[source_sheet_name]
        
        # Обрезаем имя листа до 31 символа (лимит Excel)
        safe_sheet_name = new_sheet_name[:31]
        
        # Создаем новый лист
        new_ws = self.workbook.create_sheet(title=safe_sheet_name)
        
        # Копируем все ячейки вручную с прямым доступом
        from openpyxl.utils import get_column_letter
        
        # Определяем границы - используем фиксированные значения для шаблона
        max_row = 501  # Фиксированное количество строк в шаблоне
        max_col = 21   # Фиксированное количество колонок (A-U)
        
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = source_ws.cell(row=row_idx, column=col_idx)
                coord = f"{get_column_letter(col_idx)}{row_idx}"
                new_cell = new_ws[coord]
                
                # Копируем значение или формулу
                if cell.value is not None:
                    new_cell.value = cell.value
                
                # Копируем стили
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy.copy(cell.alignment)
        
        # Копируем размеры колонок
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
            new_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        
        # Копируем размеры строк
        for row_num, row_dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height
            new_ws.row_dimensions[row_num].hidden = row_dim.hidden
        
        # Копируем объединенные ячейки
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))
    
    def write_data_to_template_sheet(
        self,
        sheet_name: str,
        data_rows: List[List],
        start_data_row: int = 11,
        column_mapping: Optional[Dict[int, int]] = None
    ) -> int:
        """
        Записывает данные из TXT в лист-шаблон, сохраняя формулы.
        
        Данные записываются в столбцы A-H (№, f MHz, P dBm, IG mA, ID A, Gain dB, КПД %, Pвых W).
        Формулы в столбцах P-U автоматически пересчитываются Excel.
        
        Args:
            sheet_name: Имя листа для записи
            data_rows: Строки данных из TXT
            start_data_row: Начальная строка для данных (по умолчанию 11)
            column_mapping: Маппинг колонок {исходная_колонка: целевая_колонка}
                           По умолчанию: {0:0, 1:1, 2:2, 3:3, 4:4, 5:5, 6:6, 7:7}
            
        Returns:
            Количество записанных строк
        """
        if self.workbook is None:
            self.load_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
        
        ws = self.workbook[sheet_name]
        
        if column_mapping is None:
            column_mapping = {i: i for i in range(8)}
        
        rows_written = 0
        for i, row_data in enumerate(data_rows):
            target_row = start_data_row + i
            for src_col, tgt_col in column_mapping.items():
                if src_col < len(row_data):
                    value = row_data[src_col]
                    # Преобразуем строковые значения в числа где возможно
                    cell_value = self._convert_value(value)
                    ws.cell(row=target_row, column=tgt_col + 1, value=cell_value)
            rows_written += 1

        first_empty_row = start_data_row + rows_written
        if first_empty_row <= ws.max_row:
            target_columns = sorted({tgt_col + 1 for tgt_col in column_mapping.values()})
            self._clear_columns_range(ws, first_empty_row, ws.max_row, target_columns)
        
        return rows_written
    
    def write_formula_columns(
        self,
        sheet_name: str,
        start_row: int = 11,
        num_rows: int = 86
    ) -> int:
        """
        Записывает формулы в колонки P-U для использования в графиках.
        
        Формулы:
          P - Pвых дБм: =C{row}+J{row}
          Q - Ig мА: =D{row}-L{row}
          R - Id А: =(E{row}-M{row})*K{row}
          S - Gain dB: =F{row}+J{row}
          T - КПД %: =((U{row}-O{row})/(N{row}*R{row}))*100
          U - Pвых Вт: =(10^(P{row}/10))/1000
        
        Args:
            sheet_name: Имя листа
            start_row: Начальная строка для формул
            num_rows: Количество строк
            
        Returns:
            Количество записанных строк
        """
        if self.workbook is None:
            self.load_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
        
        ws = self.workbook[sheet_name]
        
        for i in range(num_rows):
            row = start_row + i
            ws.cell(row=row, column=16, value=f"=C{row}+J{row}")  # P - Pвых дБм
            ws.cell(row=row, column=17, value=f"=D{row}-L{row}")  # Q - Ig мА
            ws.cell(row=row, column=18, value=f"=(E{row}-M{row})*K{row}")  # R - Id А
            ws.cell(row=row, column=19, value=f"=F{row}+J{row}")  # S - Gain dB
            ws.cell(row=row, column=20, value=f"=((U{row}-O{row})/(N{row}*R{row}))*100")  # T - КПД %
            ws.cell(row=row, column=21, value=f"=(10^(P{row}/10))/1000")  # U - Pвых Вт
        
        return num_rows
    
    def write_metadata_to_template_sheet(
        self,
        sheet_name: str,
        metadata: Dict[str, str]
    ) -> None:
        """
        Записывает метаданные из TXT в ячейки шаблона (строки 1-8, колонки A-B).
        
        Маппинг метаданных (ключ TXT -> (строка, колонка_ключа, колонка_значения)):
          - 'Мощность на УБМ, dBm' -> A1:B1
          - 'Gate start, ms' -> A2:B2
          - 'Gate lenght, ms' -> A3:B3
          - 'T, ms' -> A4:B4
          - 'Точность подстройки, dBm' -> A5:B5
          - 'Iпот(покоя), mA' -> A6:B6
          - 'Uп, В' -> A7:B7
          - 'Uсм, В' -> A8:B8
        
        Args:
            sheet_name: Имя листа
            metadata: Словарь метаданных из TXT файла
        """
        if self.workbook is None:
            self.load_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
        
        ws = self.workbook[sheet_name]
        
        # Маппинг ключей метаданных на (строка, колонка_ключа, колонка_значения)
        # Ключи должны точно соответствовать тем, что в TXT файле
        metadata_mapping = {
            'Мощность на УБМ, dBm': (1, 1, 2),      # A1:B1
            'Gate start, ms': (2, 1, 2),             # A2:B2
            'Gate lenght, ms': (3, 1, 2),            # A3:B3
            'T, ms': (4, 1, 2),                      # A4:B4
            'Точность подстройки, dBm': (5, 1, 2),   # A5:B5
            'Iпот(покоя), mA': (6, 1, 2),            # A6:B6
            'Uп, В': (7, 1, 2),                      # A7:B7
            'Uсм, В': (8, 1, 2),                     # A8:B8
        }
        
        for key, (row, key_col, value_col) in metadata_mapping.items():
            if key in metadata:
                # Записываем ключ в колонку A
                ws.cell(row=row, column=key_col, value=key)
                # Записываем значение в колонку B
                value = self._convert_value(metadata[key])
                ws.cell(row=row, column=value_col, value=value)
    
    def add_chart_series(
        self,
        charts_sheet_name: str,
        new_data_sheet_name: str,
        series_name: str,
        num_data_rows: int,
    ) -> None:
        """
        Добавляет новую серию данных на лист Графики.
        
        Находит первый свободный слот серии в строке 5 в диапазоне C-J (колонки 3-10),
        записывает туда имя серии и добавляет формулы для данных.
        
        Структура листа Графики:
          - Строка 4: заголовки (Частота Гц, Частота ГГц, Pвых Вт, Pвых дБм, Id А, Ig мА, Ку дБ, КПД %)
          - Строка 5: имена серий (C-J для первой группы, остальные группы ссылаются на них)
          - Строки 6+: данные с формулами
        
          Колонки:
          A: Частота, Гц (из B листа данных)
          B: Частота, ГГц (A/1000000000)
          C-J: Pвых, Вт (из U листа данных) - серии 1-8
          K-R: Pвых, дБм (из P листа данных) - серии 1-8
          S-Z: Id, А (из R листа данных) - серии 1-8
          AA-AH: Ig, мА (из Q листа данных) - серии 1-8
          AI-AP: Ку, дБ (из S листа данных) - серии 1-8
          AQ-AX: КПД, % (из T листа данных) - серии 1-8
        
        Максимум 8 серий на листе.
        
        Args:
            charts_sheet_name: Имя листа с графиками (например, 'Графики')
            new_data_sheet_name: Имя нового листа с данными
            series_name: Имя серии для легенды
            num_data_rows: Количество строк данных
        """
        if self.workbook is None:
            self.load_workbook()
        
        if charts_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{charts_sheet_name}' не найден")
        
        charts_ws = self.workbook[charts_sheet_name]
        
        # Ищем первый свободный слот серии (0-7)
        # Каждая серия занимает одну колонку в каждой из 6 групп по 8 колонок:
        #   Группа 0 (Pвых Вт): C-J (3-10)
        #   Группа 1 (Pвых дБм): K-R (11-18)
        #   Группа 2 (Id А): S-Z (19-26)
        #   Группа 3 (Ig мА): AA-AH (27-34)
        #   Группа 4 (Ку дБ): AI-AP (35-42)
        #   Группа 5 (КПД %): AQ-AX (43-50)
        # Имена серий хранятся в первой группе (C-J), остальные группы ссылаются на них
        series_index = None
        for idx in range(8):  # Максимум 8 серий
            col = 3 + idx  # C-J
            cell = charts_ws.cell(row=5, column=col)
            cell_value = cell.value
            # Проверяем, свободна ли ячейка:
            # - пустая (None или пустая строка)
            # - содержит формулу с #REF! (битая ссылка)
            is_free = False
            if cell_value is None or str(cell_value).strip() == '':
                is_free = True
            elif isinstance(cell_value, str) and cell_value.startswith('=') and 'REF' in str(cell_value):
                is_free = True
            
            if is_free:
                series_index = idx
                break
        
        if series_index is None:
            print(f"  Предупреждение: Нет свободных слотов для добавления новой серии (максимум 8 серий, C-J)")
            print(f"  Серия '{series_name}' не будет добавлена на лист '{charts_sheet_name}'")
            return  # Пропускаем добавление серии
        
        # Записываем имя серии в строку 5 первой группы
        selected_col = 3 + series_index
        charts_ws.cell(row=5, column=selected_col, value=series_name)
        
        # Вычисляем смещение (offset) относительно первой колонки C (3)
        offset = series_index
        
        # Маппинг колонок данных на колонки на листе Графики
        # Базовые колонки для каждого типа данных:
        #   Pвых Вт: колонки C-J (3-10) -> base_col = 3 + offset
        #   Pвых дБм: колонки K-R (11-18) -> base_col = 11 + offset
        #   Id А: колонки S-Z (19-26) -> base_col = 19 + offset
        #   Ig мА: колонки AA-AH (27-34) -> base_col = 27 + offset
        #   Gain dB: колонки AI-AP (35-42) -> base_col = 35 + offset
        #   КПД %: колонки AQ-AX (43-50) -> base_col = 43 + offset
        
        column_mapping = {
            'Pout_W': 3 + offset,
            'Pout_dBm': 11 + offset,
            'Id_A': 19 + offset,
            'Ig_mA': 27 + offset,
            'Gain_dB': 35 + offset,
            'Efficiency': 43 + offset,
        }
        
        # Записываем формулы для каждой строки данных
        for i in range(num_data_rows):
            source_row = 11 + i  # Строка в листе данных
            target_row = 6 + i   # Строка на листе Графики
            
            # Частота Гц (B в листе данных) - колонка A
            charts_ws.cell(row=target_row, column=1,
                          value=f"='{new_data_sheet_name}'!B{source_row}")
            
            # Частота ГГц (A/1000000000) - колонка B
            charts_ws.cell(row=target_row, column=2,
                          value=f"=A{target_row}/1000000000")
            
            # Pвых Вт (U в листе данных)
            charts_ws.cell(row=target_row, column=column_mapping['Pout_W'],
                          value=f"='{new_data_sheet_name}'!U{source_row}")
            
            # Pвых дБм (P в листе данных)
            charts_ws.cell(row=target_row, column=column_mapping['Pout_dBm'],
                          value=f"='{new_data_sheet_name}'!P{source_row}")
            
            # Id А (R в листе данных)
            charts_ws.cell(row=target_row, column=column_mapping['Id_A'],
                          value=f"='{new_data_sheet_name}'!R{source_row}")
            
            # Ig мА (Q в листе данных)
            charts_ws.cell(row=target_row, column=column_mapping['Ig_mA'],
                          value=f"='{new_data_sheet_name}'!Q{source_row}")
            
            # Gain dB (S в листе данных)
            charts_ws.cell(row=target_row, column=column_mapping['Gain_dB'],
                          value=f"='{new_data_sheet_name}'!S{source_row}")
            
            # КПД % (T в листе данных)
            charts_ws.cell(row=target_row, column=column_mapping['Efficiency'],
                          value=f"='{new_data_sheet_name}'!T{source_row}")
        
        # Обновляем диапазоны данных в существующих графиках
        self._update_chart_ranges(charts_sheet_name, num_data_rows)
        
        return series_index
    
    def _update_chart_ranges(self, charts_sheet_name: str, num_data_rows: int) -> None:
        """
        Обновляет диапазоны данных в графиках на листе.
        
        Находит все графики на листе и обновляет диапазоны X и Y значений
        в соответствии с количеством строк данных.
        Также добавляет новые серии для каждой новой колонки данных.
        Очищает кэшированные данные, чтобы Excel пересчитал формулы.
        
        Args:
            charts_sheet_name: Имя листа с графиками
            num_data_rows: Количество строк данных
        """
        if self.workbook is None:
            self.load_workbook()
        
        charts_ws = self.workbook[charts_sheet_name]
        
        # Определяем конечную строку данных
        end_row = 6 + num_data_rows - 1  # Строка 6 - начальная
        
        # Определяем сколько серий данных уже есть в строке 5
        existing_series = []
        for col in range(3, 11):  # C-J
            cell = charts_ws.cell(row=5, column=col)
            if cell.value is not None and str(cell.value).strip() != '' and not str(cell.value).startswith('='):
                existing_series.append(col)
        
        # Базовые колонки для каждого типа графика
        chart_y_columns = [
            (3, 'C'),   # Pвых Вт
            (11, 'K'),  # Pвых дБм
            (19, 'S'),  # Id А
            (27, 'AA'), # Ig мА
            (35, 'AI'), # Ку дБ
            (43, 'AQ'), # КПД %
        ]
        
        # Проходим по всем графикам на листе
        for chart_idx, chart in enumerate(charts_ws._charts):
            if chart_idx >= len(chart_y_columns):
                break
            
            base_col, col_letter = chart_y_columns[chart_idx]
            
            # Сбрасываем масштаб осей для автомасштабирования
            # Удаляем жестко заданные min/max, чтобы Excel мог автоматически определить масштаб
            if hasattr(chart.x_axis, 'scaling') and chart.x_axis.scaling:
                chart.x_axis.scaling.min = None
                chart.x_axis.scaling.max = None
            if hasattr(chart.y_axis, 'scaling') and chart.y_axis.scaling:
                chart.y_axis.scaling.min = None
                chart.y_axis.scaling.max = None
            
            # Обновляем существующие серии и очищаем кэш
            for series_idx, series in enumerate(chart.series):
                # Обновляем X значения (Частота ГГц - колонка B)
                if hasattr(series, 'xVal') and series.xVal:
                    if hasattr(series.xVal, 'numRef') and series.xVal.numRef:
                        series.xVal.numRef.f = "Графики!$B$6:$B$" + str(end_row)
                        # Очищаем кэшированные данные
                        if hasattr(series.xVal.numRef, 'numCache') and series.xVal.numRef.numCache:
                            series.xVal.numRef.numCache.ptCount = None
                            series.xVal.numRef.numCache.pt = []
                
                # Обновляем Y значения и очищаем кэш
                if hasattr(series, 'yVal') and series.yVal:
                    if hasattr(series.yVal, 'numRef') and series.yVal.numRef:
                        current_ref = series.yVal.numRef.f
                        # Извлекаем букву колонки из текущей ссылки
                        match = re.search(r'\$([A-Z]+)\$', current_ref)
                        if match:
                            series_col_letter = match.group(1)
                            series.yVal.numRef.f = "Графики!$" + series_col_letter + "$6:$" + series_col_letter + "$" + str(end_row)
                            # Очищаем кэшированные данные
                            if hasattr(series.yVal.numRef, 'numCache') and series.yVal.numRef.numCache:
                                series.yVal.numRef.numCache.ptCount = None
                                series.yVal.numRef.numCache.pt = []
            
            # Добавляем новые серии для каждой новой колонки данных
            # Каждая серия соответствует колонке в диапазоне base_col до base_col+7
            for series_idx, data_col in enumerate(range(base_col, base_col + len(existing_series))):
                if series_idx >= len(chart.series):
                    # Создаем новую серию
                    from openpyxl.chart.series import XYSeries
                    from openpyxl.chart.data_source import NumDataSource, AxDataSource, NumRef, StrRef, StrData, StrVal
                    from openpyxl.chart.marker import Marker
                    from openpyxl.drawing.colors import ColorChoice
                    from openpyxl.chart.shapes import GraphicalProperties
                    from openpyxl.drawing.line import LineProperties
                    
                    col_letter_for_series = openpyxl.utils.get_column_letter(data_col)
                    
                    x_ref = NumRef("Графики!$B$6:$B$" + str(end_row))
                    y_ref = NumRef("Графики!$" + col_letter_for_series + "$6:$" + col_letter_for_series + "$" + str(end_row))
                    
                    # Определяем цвет для серии
                    colors = ['0070C0', 'FF0000', '00B050', 'FFC000', '7030A0', '00B0F0', 'FF6600', '92D050']
                    color = colors[series_idx % len(colors)]
                    
                    series = XYSeries(
                        idx=series_idx,
                        tx=None,
                        spPr=GraphicalProperties(
                            ln=LineProperties(
                                noFill=False,
                                solidFill=ColorChoice(srgbClr=color)
                            )
                        ),
                        xVal=AxDataSource(numRef=x_ref),
                        yVal=NumDataSource(numRef=y_ref),
                        smooth=True,
                    )
                    chart.series.append(series)
    
    def create_new_chart(
        self,
        charts_sheet_name: str,
        chart_title: str,
        y_column_name: str,
        y_column_letter: str,
        y_axis_title: str,
        position_col: int = 1,
        position_row: int = 1,
    ) -> None:
        """
        Создает новый график на листе Графики.
        
        Args:
            charts_sheet_name: Имя листа с графиками
            chart_title: Заголовок графика
            y_column_name: Имя типа данных для Y оси (например, 'Pout_W')
            y_column_letter: Буква колонки для Y значений (например, 'C')
            y_axis_title: Заголовок оси Y
            position_col: Начальная колонка для размещения графика
            position_row: Начальная строка для размещения графика
        """
        from openpyxl.chart import ScatterChart, Reference
        from openpyxl.chart.series import XYSeries
        from openpyxl.chart.data_source import NumDataSource, AxDataSource, NumRef
        from openpyxl.chart.label import DataLabelList
        
        if self.workbook is None:
            self.load_workbook()
        
        charts_ws = self.workbook[charts_sheet_name]
        
        # Создаем новый ScatterChart
        chart = ScatterChart()
        chart.title = chart_title
        chart.style = 10
        chart.graphical_properties = None
        
        # Настройка оси X
        chart.x_axis.title = "Частота, ГГц"
        chart.x_axis.scaling.min = 0
        # Не устанавливаем max жестко - будет автомасштабирование
        
        # Настройка оси Y
        chart.y_axis.title = y_axis_title
        chart.y_axis.scaling.min = 0
        # Не устанавливаем max жестко - будет автомасштабирование
        
        # Настройка легенды
        chart.legend.position = "b"
        
        # Добавляем серию данных (будет заполнена при обновлении)
        # X: колонка B (Частота ГГц)
        # Y: указанная колонка
        x_ref = NumRef("Графики!$B$6:$B$91")
        y_ref = NumRef(f"Графики!${y_column_letter}$6:${y_column_letter}$91")
        
        series = XYSeries(
            idx=0,
            tx=None,
            spPr=None,
            xVal=AxDataSource(numRef=x_ref),
            yVal=NumDataSource(numRef=y_ref),
            smooth=True,
            marker=None,
        )
        chart.series.append(series)
        
        # Размещаем график
        from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
        from openpyxl.drawing.spreadsheet_drawing import AnchorClientData
        
        anchor = TwoCellAnchor(
            _from=AnchorMarker(col=position_col, row=position_row),
            to=AnchorMarker(col=position_col + 5, row=position_row + 12),
            clientData=AnchorClientData()
        )
        anchor.sp = chart
        
        charts_ws.add_chart(chart, anchor)
    
    def _convert_value(self, value):
        """Преобразует строковое значение в число где возможно."""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return value
        
        str_value = str(value).strip().replace(',', '.')
        
        try:
            # Пробуем преобразовать в int
            if '.' not in str_value:
                return int(str_value)
        except (ValueError, TypeError):
            pass
        
        try:
            # Пробуем преобразовать в float
            return float(str_value)
        except (ValueError, TypeError):
            return value
    
    def _clear_columns_range(self, worksheet, start_row: int, end_row: int, columns) -> None:
        """Очищает заданные колонки в диапазоне строк, не затрагивая оформление листа."""
        for row_idx in range(start_row, end_row + 1):
            for col_idx in columns:
                worksheet.cell(row=row_idx, column=col_idx).value = None

    def write_s2p_data_to_template_sheet(
        self,
        sheet_name: str,
        data_rows: List[List],
        start_data_row: int = 3
    ) -> int:
        """
        Записывает данные из S2P файла в лист-шаблон.
        
        Данные S2P записываются в столбцы A-E:
          A - Freq, GHz
          B - S11M (дБ)
          C - S21M (дБ)
          D - S12M (дБ)
          E - S22M (дБ)
        
        Args:
            sheet_name: Имя листа для записи
            data_rows: Строки данных из S2P файла
            start_data_row: Начальная строка для данных (по умолчанию 3)
            
        Returns:
            Количество записанных строк
        """
        if self.workbook is None:
            self.load_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
        
        ws = self.workbook[sheet_name]
        
        rows_written = 0
        for i, row_data in enumerate(data_rows):
            target_row = start_data_row + i
            # Записываем данные в колонки A-E
            for col_idx, value in enumerate(row_data):
                if col_idx < 5:  # Только 5 колонок (A-E)
                    cell_value = self._convert_value(value)
                    ws.cell(row=target_row, column=col_idx + 1, value=cell_value)
            rows_written += 1
        
        return rows_written
