"""Internal service layer for AutoGraph jobs."""

from contextlib import redirect_stdout
from dataclasses import dataclass, field
import os
import shutil
from typing import Callable, List, Optional, Sequence, Tuple

from autograph_service.excel_handler import ExcelHandler
from autograph_service.s2p_excel_handler import S2PExcelHandler
from autograph_service.s2p_parser import S2PParser
from autograph_service.txt_parser import TxtFileParser


DATA_KIND_TXT = "txt"
DATA_KIND_S2P = "s2p"
SOURCE_KIND_FILE = "single_file"
SOURCE_KIND_DIRECTORY = "directory"
DEFAULT_TEMPLATE_SHEET = "№1"
DEFAULT_CHARTS_SHEET = "Графики"

Reporter = Optional[Callable[[str], None]]


class JobValidationError(ValueError):
    """Raised when a job configuration is invalid."""


@dataclass
class JobConfig:
    """Configuration for a single AutoGraph run."""

    data_kind: str
    source_kind: str
    input_path: str
    excel_path: str
    template_mode: bool = False
    sheet_name: Optional[str] = None
    output_path: Optional[str] = None
    template_sheet: str = DEFAULT_TEMPLATE_SHEET
    charts_sheet: str = DEFAULT_CHARTS_SHEET
    charts_sheet_template: str = DEFAULT_CHARTS_SHEET
    use_template_charts: bool = True
    create_chart_specs: List[Tuple[str, str, str]] = field(default_factory=list)


@dataclass
class JobResult:
    """Summary of a completed AutoGraph run."""

    output_paths: List[str]
    processed_files: int


def run_job(config: JobConfig, reporter: Reporter = None) -> JobResult:
    """Execute an AutoGraph job using the provided configuration."""
    config = validate_job_config(config)

    with redirect_stdout(_ReporterStream(reporter)):
        file_type = "S2P" if config.data_kind == DATA_KIND_S2P else "TXT"
        _report(reporter, f"AutoGraph Service - Перенос данных из {file_type} в Excel")
        _report(reporter, "=" * 50)
        _report(reporter, f"Excel файл: {config.excel_path}")
        _report(reporter, f"Режим шаблона: {'Да' if config.template_mode else 'Нет'}")
        if config.template_mode:
            _report(reporter, f"Шаблон листа: {config.template_sheet}")
            if config.data_kind == DATA_KIND_TXT:
                _report(reporter, f"Лист графиков: {config.charts_sheet}")
            else:
                _report(
                    reporter,
                    f"Графики из шаблона: {'Да' if config.use_template_charts else 'Нет'}",
                )
                _report(reporter, f"Лист графиков в шаблоне: {config.charts_sheet_template}")
        if config.output_path and config.source_kind == SOURCE_KIND_FILE:
            _report(reporter, f"Выходной файл: {config.output_path}")

        if config.data_kind == DATA_KIND_S2P:
            if config.source_kind == SOURCE_KIND_DIRECTORY:
                return _process_s2p_directory(config, reporter)
            return _process_s2p_single(config, reporter)

        if config.source_kind == SOURCE_KIND_DIRECTORY:
            return _process_txt_directory(config, reporter)
        return _process_txt_single(config, reporter)


def validate_job_config(config: JobConfig) -> JobConfig:
    """Validate a job configuration and normalize obvious input issues."""
    normalized = JobConfig(
        data_kind=(config.data_kind or "").strip().lower(),
        source_kind=(config.source_kind or "").strip().lower(),
        input_path=(config.input_path or "").strip(),
        excel_path=(config.excel_path or "").strip(),
        template_mode=bool(config.template_mode),
        sheet_name=_normalize_optional(config.sheet_name),
        output_path=_normalize_optional(config.output_path),
        template_sheet=(config.template_sheet or DEFAULT_TEMPLATE_SHEET).strip(),
        charts_sheet=(config.charts_sheet or DEFAULT_CHARTS_SHEET).strip(),
        charts_sheet_template=(config.charts_sheet_template or DEFAULT_CHARTS_SHEET).strip(),
        use_template_charts=bool(config.use_template_charts),
        create_chart_specs=list(config.create_chart_specs or []),
    )

    if normalized.data_kind not in {DATA_KIND_TXT, DATA_KIND_S2P}:
        raise JobValidationError("Нужно выбрать тип входных данных: TXT или S2P.")
    if normalized.source_kind not in {SOURCE_KIND_FILE, SOURCE_KIND_DIRECTORY}:
        raise JobValidationError("Нужно выбрать источник данных: один файл или папка.")
    if not normalized.input_path:
        raise JobValidationError("Нужно указать путь к входным данным.")
    if not normalized.excel_path:
        raise JobValidationError("Нужно указать путь к Excel-файлу или шаблону.")

    if normalized.source_kind == SOURCE_KIND_FILE:
        if not os.path.isfile(normalized.input_path):
            raise JobValidationError(f"Файл данных не найден: {normalized.input_path}")
    else:
        if not os.path.isdir(normalized.input_path):
            raise JobValidationError(f"Папка с данными не найдена: {normalized.input_path}")

    if normalized.data_kind == DATA_KIND_TXT and normalized.source_kind == SOURCE_KIND_FILE:
        if not normalized.input_path.lower().endswith(".txt"):
            raise JobValidationError("Для TXT-режима нужно выбрать файл с расширением .txt.")

    if normalized.data_kind == DATA_KIND_S2P and normalized.source_kind == SOURCE_KIND_FILE:
        if not normalized.input_path.lower().endswith(".s2p"):
            raise JobValidationError("Для S2P-режима нужно выбрать файл с расширением .s2p.")

    if normalized.data_kind == DATA_KIND_S2P and not normalized.template_mode:
        raise JobValidationError("S2P поддерживается только в шаблонном режиме.")

    if normalized.template_mode:
        if not normalized.template_sheet:
            raise JobValidationError("Для шаблонного режима нужно указать имя листа-шаблона.")
        if not os.path.exists(normalized.excel_path):
            raise JobValidationError(f"Excel-шаблон не найден: {normalized.excel_path}")

    return normalized


def scan_txt_files(directory: str) -> List[str]:
    """Scan a directory and return sorted TXT files."""
    if not os.path.isdir(directory):
        raise NotADirectoryError(f"Указанный путь не является папкой: {directory}")

    txt_files = []
    for filename in os.listdir(directory):
        if filename.lower().endswith(".txt"):
            txt_files.append(os.path.join(directory, filename))

    txt_files.sort(key=lambda path: os.path.basename(path))
    return txt_files


def scan_s2p_files(directory: str) -> List[str]:
    """Scan a directory and return sorted S2P files."""
    if not os.path.isdir(directory):
        raise NotADirectoryError(f"Указанный путь не является папкой: {directory}")

    s2p_files = []
    for filename in os.listdir(directory):
        if filename.lower().endswith(".s2p"):
            s2p_files.append(os.path.join(directory, filename))

    s2p_files.sort(key=lambda path: os.path.basename(path))
    return s2p_files


def group_files(file_paths: Sequence[str], max_per_group: int = 8) -> List[List[str]]:
    """Split files into groups of up to ``max_per_group`` items."""
    groups: List[List[str]] = []
    for index in range(0, len(file_paths), max_per_group):
        groups.append(list(file_paths[index:index + max_per_group]))
    return groups


def generate_output_filename(original_excel_path: str, group_index: int, total_groups: int) -> str:
    """Generate an output workbook path for TXT batch processing."""
    directory = os.path.dirname(original_excel_path) or "."
    basename = os.path.basename(original_excel_path)
    name, ext = os.path.splitext(basename)

    if total_groups == 1:
        return original_excel_path

    return os.path.join(directory, f"{name}_part{group_index + 1}{ext}")


def generate_sheet_name(file_path: str) -> str:
    """Generate a safe Excel sheet name from a file path."""
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    for char in ["\\", "/", "?", "*", "[", "]", ":"]:
        base_name = base_name.replace(char, "_")
    return base_name[:31]


def generate_unique_sheet_name(existing_sheets: Sequence[str], base_name: str) -> str:
    """Generate a unique sheet name within Excel length limits."""
    base_name = base_name[:31]
    if base_name not in existing_sheets:
        return base_name

    max_base_len = 31 - 3
    counter = 1
    while counter < 100:
        short_base = base_name[:max_base_len]
        new_name = f"{short_base}_{counter}"
        if new_name not in existing_sheets:
            return new_name[:31]
        counter += 1
    return base_name


def generate_output_excel_name(s2p_path: str, output_path: Optional[str] = None) -> str:
    """Generate a target workbook path for S2P output."""
    if output_path:
        result = output_path
    else:
        base_name = os.path.splitext(os.path.basename(s2p_path))[0]
        directory = os.path.dirname(s2p_path) or "."
        result = os.path.join(directory, f"{base_name}.xlsx")

    if not result.lower().endswith(".xlsx"):
        result += ".xlsx"
    return result


def resolve_s2p_chart_sheet(
    template_handler: S2PExcelHandler,
    template_sheet_name: str,
    preferred_sheet_name: str,
) -> Tuple[Optional[str], bool]:
    """Resolve the chart sheet source for S2P template processing."""
    template_ws = template_handler.workbook[template_sheet_name]
    if getattr(template_ws, "_charts", None):
        return template_sheet_name, True

    sheet_names = template_handler.get_sheet_names()
    chart_sheets = [
        name
        for name in sheet_names
        if name != template_sheet_name and getattr(template_handler.workbook[name], "_charts", None)
    ]

    if preferred_sheet_name and preferred_sheet_name != DEFAULT_CHARTS_SHEET and preferred_sheet_name in chart_sheets:
        return preferred_sheet_name, False

    s2p_chart_sheets = [name for name in chart_sheets if "s2p" in name.lower()]
    if s2p_chart_sheets:
        return s2p_chart_sheets[0], False
    if preferred_sheet_name in chart_sheets:
        return preferred_sheet_name, False
    if len(chart_sheets) == 1:
        return chart_sheets[0], False
    if chart_sheets:
        return chart_sheets[0], False
    return None, False


def _process_txt_single(config: JobConfig, reporter: Reporter) -> JobResult:
    file_path = config.input_path
    sheet_name = config.sheet_name or generate_sheet_name(file_path)

    _report(reporter, "Режим: Одиночный файл")
    _report(reporter, f"TXT файл: {file_path}")
    if not config.template_mode:
        _report(reporter, f"Имя листа: {sheet_name}")
    _report(reporter, "=" * 50)

    handler = ExcelHandler(config.excel_path)
    handler.load_workbook()
    _process_txt_file(config, file_path, handler, reporter, is_first_file=True)

    output_path = config.output_path or None
    handler.save(output_path)
    save_path = output_path or config.excel_path
    _report(reporter, f"Данные сохранены в: {save_path}")
    _report(reporter, "Готово!")
    return JobResult(output_paths=[save_path], processed_files=1)


def _process_txt_directory(config: JobConfig, reporter: Reporter) -> JobResult:
    _report(reporter, "Режим: Пакетная обработка папки")
    _report(reporter, f"Папка с TXT файлами: {config.input_path}")

    files = scan_txt_files(config.input_path)
    if not files:
        raise FileNotFoundError(f"В папке '{config.input_path}' не найдено TXT файлов")

    file_groups = group_files(files, max_per_group=8)
    total_groups = len(file_groups)

    _report(reporter, f"Найдено TXT файлов: {len(files)}")
    _report(reporter, f"Групп файлов (по 8 максимум): {total_groups}")
    _report(reporter, "=" * 50)

    output_paths: List[str] = []
    processed_files = 0

    for group_idx, current_group in enumerate(file_groups):
        _report(reporter, "=" * 50)
        _report(reporter, f"Группа {group_idx + 1}/{total_groups} ({len(current_group)} файлов)")

        output_path = generate_output_filename(config.excel_path, group_idx, total_groups)
        if group_idx == 0:
            current_excel_path = config.excel_path
        else:
            prev_output = generate_output_filename(config.excel_path, group_idx - 1, total_groups)
            shutil.copy2(prev_output, output_path)
            current_excel_path = output_path
            _report(reporter, f"Создана копия Excel файла: {output_path}")

        _report(reporter, f"Excel файл: {current_excel_path}")
        _report(reporter, f"Сохранение в: {output_path}")

        handler = ExcelHandler(current_excel_path)
        handler.load_workbook()

        for file_idx, file_path in enumerate(current_group):
            is_first_file = group_idx == 0 and file_idx == 0
            _process_txt_file(config, file_path, handler, reporter, is_first_file=is_first_file)
            processed_files += 1

        handler.save(output_path)
        output_paths.append(output_path)
        _report(reporter, f"Сохранено в: {output_path}")
        _report(reporter, f"Листы в файле: {handler.get_sheet_names()}")

    _report(reporter, "=" * 50)
    _report(reporter, "Пакетная обработка завершена!")
    _report(reporter, f"Создано файлов: {len(output_paths)}")
    for path in output_paths:
        _report(reporter, f"  - {path}")
    _report(reporter, "Готово!")
    return JobResult(output_paths=output_paths, processed_files=processed_files)


def _process_txt_file(
    config: JobConfig,
    file_path: str,
    excel_handler: ExcelHandler,
    reporter: Reporter,
    is_first_file: bool,
) -> None:
    sheet_name = config.sheet_name or generate_sheet_name(file_path)

    _report(reporter, f"Обработка файла: {file_path}")
    _report(reporter, f"Тип файла: TXT")
    _report(reporter, f"Целевой лист: {sheet_name}")

    parser = TxtFileParser()
    measurement_data = parser.parse_file(file_path)
    headers = measurement_data.headers
    rows = measurement_data.rows
    metadata = measurement_data.metadata

    _report(reporter, f"Найдено строк данных: {len(rows)}")
    _report(reporter, f"Количество колонок: {len(headers)}")
    if metadata:
        _report(reporter, f"Метаданные: {list(metadata.keys())}")

    if config.template_mode:
        if config.template_sheet not in excel_handler.get_sheet_names():
            raise ValueError(
                f"Шаблон листа '{config.template_sheet}' не найден. "
                f"Доступные листы: {excel_handler.get_sheet_names()}"
            )

        new_sheet_name = generate_unique_sheet_name(excel_handler.get_sheet_names(), sheet_name)
        _report(reporter, f"Новый лист: {new_sheet_name}")

        excel_handler.copy_template_sheet(
            source_sheet_name=config.template_sheet,
            new_sheet_name=new_sheet_name,
        )
        _report(reporter, f"Шаблон '{config.template_sheet}' скопирован в '{new_sheet_name}'")

        rows_written = excel_handler.write_data_to_template_sheet(
            sheet_name=new_sheet_name,
            data_rows=rows,
            start_data_row=11,
        )
        _report(reporter, f"Записано в A-H: {rows_written} строк")

        excel_handler.write_data_columns_pu(
            sheet_name=new_sheet_name,
            data_rows=rows,
            start_row=11,
        )
        _report(reporter, f"Данные записаны в P-U: {rows_written} строк")

        excel_handler.write_metadata_to_template_sheet(
            sheet_name=new_sheet_name,
            metadata=metadata,
        )
        _report(reporter, "Метаданные записаны в строку 1")

        if config.charts_sheet in excel_handler.get_sheet_names():
            result = excel_handler.add_chart_series(
                charts_sheet_name=config.charts_sheet,
                new_data_sheet_name=new_sheet_name,
                series_name=new_sheet_name,
                num_data_rows=rows_written,
            )
            if result is not None:
                _report(
                    reporter,
                    f"Серия '{new_sheet_name}' добавлена на лист "
                    f"'{config.charts_sheet}' (слот {result + 1})",
                )

            if config.create_chart_specs and is_first_file:
                _report(reporter, "Создание новых графиков...")
                chart_positions = [
                    (1, 92),
                    (7, 92),
                    (13, 92),
                    (19, 92),
                    (25, 92),
                    (31, 92),
                    (37, 92),
                    (43, 92),
                ]
                for index, (y_col, col_letter, y_title) in enumerate(config.create_chart_specs):
                    if index >= len(chart_positions):
                        break
                    pos_col, pos_row = chart_positions[index]
                    excel_handler.create_new_chart(
                        charts_sheet_name=config.charts_sheet,
                        chart_title=y_title,
                        y_column_name=y_col,
                        y_column_letter=col_letter,
                        y_axis_title=y_title,
                        position_col=pos_col,
                        position_row=pos_row,
                    )
                    _report(reporter, f"Создан график '{y_title}' (колонка {col_letter})")
        else:
            _report(
                reporter,
                f"Предупреждение: Лист '{config.charts_sheet}' не найден, графики не обновлены",
            )
        return

    if sheet_name in excel_handler.get_sheet_names():
        _report(reporter, f"Лист '{sheet_name}' будет перезаписан")

    excel_handler.overwrite_sheet(sheet_name=sheet_name, headers=headers, rows=rows)


def _process_s2p_single(config: JobConfig, reporter: Reporter) -> JobResult:
    output_path = generate_output_excel_name(config.input_path, config.output_path)
    _report(reporter, "Режим: Одиночный файл")
    _report(reporter, f"S2P файл: {config.input_path}")
    _report(reporter, "=" * 50)

    parser = S2PParser()
    data = parser.parse_file(config.input_path)
    _report(reporter, f"Найдено строк данных: {len(data.rows)}")
    if data.rows:
        _report(reporter, f"Частота: {data.rows[0][0]} - {data.rows[-1][0]} GHz")

    template_handler = S2PExcelHandler(config.excel_path)
    template_handler.load_workbook()
    if config.template_sheet not in template_handler.get_sheet_names():
        raise ValueError(
            f"Шаблон листа '{config.template_sheet}' не найден. "
            f"Доступные листы: {template_handler.get_sheet_names()}"
        )

    import openpyxl

    new_workbook = openpyxl.Workbook()
    if "Sheet" in new_workbook.sheetnames:
        del new_workbook["Sheet"]

    sheet_name = (config.sheet_name or config.template_sheet)[:31]
    _report(reporter, f"Создание нового Excel файла: {output_path}")
    _report(reporter, f"Копирование листа '{config.template_sheet}' -> '{sheet_name}'")
    template_handler.copy_template_sheet_to_new_workbook(
        source_sheet_name=config.template_sheet,
        new_workbook=new_workbook,
        new_sheet_name=sheet_name,
    )

    use_charts = config.use_template_charts
    charts_sheet_name = sheet_name
    if use_charts:
        chart_source_sheet, charts_embedded = resolve_s2p_chart_sheet(
            template_handler,
            config.template_sheet,
            config.charts_sheet_template,
        )
        if chart_source_sheet is None:
            use_charts = False
            _report(reporter, "Предупреждение: лист с графиками для S2P не найден")
        elif not charts_embedded:
            charts_sheet_name = generate_unique_sheet_name(new_workbook.sheetnames, chart_source_sheet)
            _report(reporter, f"Копирование листа графиков '{chart_source_sheet}' -> '{charts_sheet_name}'")
            template_handler.copy_template_sheet_to_new_workbook(
                source_sheet_name=chart_source_sheet,
                new_workbook=new_workbook,
                new_sheet_name=charts_sheet_name,
            )

    handler = S2PExcelHandler(output_path)
    handler.workbook = new_workbook

    rows_written = handler.write_s2p_data(sheet_name, data.rows, start_data_row=2)
    _report(reporter, f"Записано S2P данных: {rows_written} строк")

    if use_charts:
        analysis = handler.analyze_s2p_data(data.rows)
        handler.update_chart_data_ranges(
            charts_sheet_name,
            sheet_name,
            len(data.rows),
            start_data_row=2,
            s2p_mode=True,
        )
        handler.auto_scale_charts(charts_sheet_name, analysis)
        _report(reporter, "Диапазоны данных в графиках обновлены")
        _report(reporter, "Графики автоматически масштабированы")

    handler.save(output_path)
    _report(reporter, f"Сохранено в: {output_path}")
    _report(reporter, f"Листы в файле: {handler.get_sheet_names()}")
    _report(reporter, "Готово!")
    return JobResult(output_paths=[output_path], processed_files=1)


def _process_s2p_directory(config: JobConfig, reporter: Reporter) -> JobResult:
    s2p_files = scan_s2p_files(config.input_path)
    if not s2p_files:
        raise FileNotFoundError(f"В папке '{config.input_path}' не найдено S2P файлов")

    _report(reporter, f"Найдено S2P файлов: {len(s2p_files)}")
    _report(reporter, "=" * 50)

    template_handler = S2PExcelHandler(config.excel_path)
    template_handler.load_workbook()
    if config.template_sheet not in template_handler.get_sheet_names():
        raise ValueError(
            f"Шаблон листа '{config.template_sheet}' не найден. "
            f"Доступные листы: {template_handler.get_sheet_names()}"
        )

    import openpyxl

    output_paths: List[str] = []
    for index, s2p_file in enumerate(s2p_files):
        _report(reporter, "=" * 50)
        _report(reporter, f"Файл {index + 1}/{len(s2p_files)}")

        output_path = generate_output_excel_name(s2p_file, config.output_path)
        if len(s2p_files) > 1:
            base, ext = os.path.splitext(output_path)
            output_path = f"{base}_{index + 1}{ext}"

        _report(reporter, f"Обработка: {s2p_file}")
        _report(reporter, f"Выходной файл: {output_path}")

        data = S2PParser().parse_file(s2p_file)
        new_workbook = openpyxl.Workbook()
        if "Sheet" in new_workbook.sheetnames:
            del new_workbook["Sheet"]

        base_sheet_name = config.sheet_name or os.path.splitext(os.path.basename(s2p_file))[0]
        sheet_name = generate_unique_sheet_name([], base_sheet_name)
        _report(reporter, f"Копирование листа '{config.template_sheet}' -> '{sheet_name}'")
        template_handler.copy_template_sheet_to_new_workbook(
            source_sheet_name=config.template_sheet,
            new_workbook=new_workbook,
            new_sheet_name=sheet_name,
        )

        use_charts = config.use_template_charts
        charts_sheet_name = sheet_name
        if use_charts:
            chart_source_sheet, charts_embedded = resolve_s2p_chart_sheet(
                template_handler,
                config.template_sheet,
                config.charts_sheet_template,
            )
            if chart_source_sheet is None:
                use_charts = False
                _report(reporter, "Предупреждение: лист с графиками для S2P не найден")
            elif not charts_embedded:
                charts_sheet_name = generate_unique_sheet_name(new_workbook.sheetnames, chart_source_sheet)
                _report(reporter, f"Копирование листа графиков '{chart_source_sheet}' -> '{charts_sheet_name}'")
                template_handler.copy_template_sheet_to_new_workbook(
                    source_sheet_name=chart_source_sheet,
                    new_workbook=new_workbook,
                    new_sheet_name=charts_sheet_name,
                )

        handler = S2PExcelHandler(output_path)
        handler.workbook = new_workbook
        rows_written = handler.write_s2p_data(sheet_name, data.rows, start_data_row=2)
        _report(reporter, f"Лист: {sheet_name}, строк: {rows_written}")

        if use_charts:
            analysis = handler.analyze_s2p_data(data.rows)
            handler.update_chart_data_ranges(
                charts_sheet_name,
                sheet_name,
                len(data.rows),
                start_data_row=2,
                s2p_mode=True,
            )
            handler.auto_scale_charts(charts_sheet_name, analysis)
            _report(reporter, "Графики автоматически масштабированы")

        handler.save(output_path)
        output_paths.append(output_path)
        _report(reporter, f"Сохранено в: {output_path}")
        _report(reporter, f"Листы: {handler.get_sheet_names()}")

    _report(reporter, "=" * 50)
    _report(reporter, "Пакетная обработка завершена!")
    _report(reporter, f"Создано файлов: {len(output_paths)}")
    return JobResult(output_paths=output_paths, processed_files=len(s2p_files))


def _normalize_optional(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    normalized = value.strip()
    return normalized or None


def _report(reporter: Reporter, message: str) -> None:
    if reporter is not None:
        reporter(message)


class _ReporterStream:
    """File-like object that forwards stdout lines into the reporter callback."""

    def __init__(self, reporter: Reporter):
        self.reporter = reporter
        self.buffer = ""

    def write(self, text: str) -> int:
        if not text:
            return 0

        self.buffer += text
        while "\n" in self.buffer:
            line, self.buffer = self.buffer.split("\n", 1)
            if line.strip():
                _report(self.reporter, line)
        return len(text)

    def flush(self) -> None:
        if self.buffer.strip():
            _report(self.reporter, self.buffer.strip())
        self.buffer = ""
