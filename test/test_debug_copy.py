#!/usr/bin/env python3
"""Отладка копирования шаблона."""

import openpyxl
import copy
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('R086W2-2 различные образц Н.Р..xlsx')
source_ws = wb['№1']

print(f'source_ws.min_row: {source_ws.min_row}')
print(f'source_ws.max_row: {source_ws.max_row}')
print(f'source_ws.min_column: {source_ws.min_column}')
print(f'source_ws.max_column: {source_ws.max_column}')

# Проверяем P10 и P11 в исходном листе
print(f'P10 в №1: {source_ws["P10"].value}')
print(f'P11 в №1: {source_ws["P11"].value}')

# Проверяем через cell()
print(f'P10 через cell(10, 16): {source_ws.cell(row=10, column=16).value}')
print(f'P11 через cell(11, 16): {source_ws.cell(row=11, column=16).value}')

# Создаем новый лист
new_ws = wb.create_sheet(title='TestDebug')

min_row = source_ws.min_row or 1
max_row = source_ws.max_row or 1
min_col = source_ws.min_column or 1
max_col = source_ws.max_column or 1

print(f'Копируем: min_row={min_row}, max_row={max_row}, min_col={min_col}, max_col={max_col}')

# Копируем только строки 10-12 и колонки 15-21 для отладки
for row_idx in range(10, 13):
    for col_idx in range(15, 22):
        cell = source_ws.cell(row=row_idx, column=col_idx)
        coord = f'{get_column_letter(col_idx)}{row_idx}'
        new_cell = new_ws[coord]
        
        print(f'  Копируем {coord}: value={cell.value}')
        
        if cell.value is not None:
            new_cell.value = cell.value
        
        if cell.has_style:
            new_cell.font = copy.copy(cell.font)
            new_cell.border = copy.copy(cell.border)
            new_cell.fill = copy.copy(cell.fill)
            new_cell.number_format = cell.number_format
            new_cell.alignment = copy.copy(cell.alignment)

print(f'\nP10 в TestDebug: {new_ws["P10"].value}')
print(f'P11 в TestDebug: {new_ws["P11"].value}')
print(f'U11 в TestDebug: {new_ws["U11"].value}')

wb.save('test_debug.xlsx')
print('\nСохранено в test_debug.xlsx')
