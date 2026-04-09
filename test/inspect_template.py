#!/usr/bin/env python3
"""Анализ структуры шаблона."""

import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('R086W2-2 различные образц Н.Р..xlsx')
ws = wb['№1']

print('=== Заголовки строки 1 ===')
for col in range(1, 22):
    value = ws.cell(row=1, column=col).value
    if value:
        print(f'  {col}: {value}')

print('\n=== Заголовки строки 10 ===')
for col in range(1, 22):
    value = ws.cell(row=10, column=col).value
    if value:
        print(f'  {col}: {value}')

print('\n=== Формулы строки 11 ===')
for col in range(16, 22):  # P-U
    cell = ws.cell(row=11, column=col)
    print(f'  {get_column_letter(col)}11: {cell.value}')

print('\n=== Значения строки 11 (данные из шаблона) ===')
for col in range(1, 9):
    value = ws.cell(row=11, column=col).value
    print(f'  {col}: {value}')

print('\n=== Метаданные (строка 1, колонки J, L, M, N, O) ===')
for col in [10, 12, 13, 14, 15]:
    value = ws.cell(row=1, column=col).value
    print(f'  {col}: {value}')
