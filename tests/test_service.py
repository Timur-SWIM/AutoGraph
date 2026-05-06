import shutil
import unittest
import uuid
from pathlib import Path

import openpyxl

from autograph_service.service import (
    DATA_KIND_S2P,
    DATA_KIND_TXT,
    DEFAULT_TEMPLATE_SHEET,
    JobConfig,
    JobValidationError,
    SOURCE_KIND_DIRECTORY,
    SOURCE_KIND_FILE,
    generate_sheet_name,
    run_job,
    validate_job_config,
)


ROOT = Path(__file__).resolve().parents[1]
GRAPH_DIR = ROOT / "graph"
EXAMPLE_S2P_DIR = ROOT / "example_s2p"
TXT_TEMPLATE = next(GRAPH_DIR.glob("R086W2-2*.xlsx"))
S2P_TEMPLATE = next(GRAPH_DIR.glob("0283*.xlsx"))
TXT_SAMPLES = sorted(GRAPH_DIR.glob("*.txt"))
S2P_SAMPLE = next(EXAMPLE_S2P_DIR.glob("*.s2p"))


class ServiceIntegrationTests(unittest.TestCase):
    def setUp(self):
        base_dir = ROOT / ".codex-run" / "test-temp"
        base_dir.mkdir(parents=True, exist_ok=True)
        self.temp_dir = base_dir / uuid.uuid4().hex
        self.temp_dir.mkdir()

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def copy_file(self, source: Path, target_name: str = None) -> Path:
        destination = self.temp_dir / (target_name or source.name)
        shutil.copy2(source, destination)
        return destination

    def test_validate_job_config_blocks_s2p_without_template(self):
        s2p_file = self.copy_file(S2P_SAMPLE)
        with self.assertRaises(JobValidationError):
            validate_job_config(
                JobConfig(
                    data_kind=DATA_KIND_S2P,
                    source_kind=SOURCE_KIND_FILE,
                    input_path=str(s2p_file),
                    excel_path=str(self.temp_dir / "result.xlsx"),
                    template_mode=False,
                )
            )

    def test_run_job_txt_single_plain_creates_output_workbook(self):
        txt_file = self.copy_file(TXT_SAMPLES[0])
        output_file = self.temp_dir / "plain-result.xlsx"

        result = run_job(
            JobConfig(
                data_kind=DATA_KIND_TXT,
                source_kind=SOURCE_KIND_FILE,
                input_path=str(txt_file),
                excel_path=str(self.temp_dir / "plain-template.xlsx"),
                template_mode=False,
                output_path=str(output_file),
            )
        )

        self.assertEqual(result.processed_files, 1)
        self.assertEqual(result.output_paths, [str(output_file)])
        self.assertTrue(output_file.exists())

        workbook = openpyxl.load_workbook(output_file)
        self.assertIn(generate_sheet_name(str(txt_file)), workbook.sheetnames)

    def test_run_job_txt_single_template_creates_output_workbook(self):
        txt_file = self.copy_file(TXT_SAMPLES[0])
        template_file = self.copy_file(TXT_TEMPLATE, "txt-template.xlsx")
        output_file = self.temp_dir / "txt-template-result.xlsx"

        result = run_job(
            JobConfig(
                data_kind=DATA_KIND_TXT,
                source_kind=SOURCE_KIND_FILE,
                input_path=str(txt_file),
                excel_path=str(template_file),
                template_mode=True,
                output_path=str(output_file),
            )
        )

        self.assertEqual(result.processed_files, 1)
        self.assertTrue(output_file.exists())

        workbook = openpyxl.load_workbook(output_file)
        self.assertIn("Графики", workbook.sheetnames)
        self.assertIn(DEFAULT_TEMPLATE_SHEET, workbook.sheetnames)
        self.assertGreater(len(workbook.sheetnames), 2)

    def test_run_job_txt_directory_template_creates_group_outputs(self):
        input_dir = self.temp_dir / "txt-batch"
        input_dir.mkdir()
        for index, source in enumerate(TXT_SAMPLES[:9], start=1):
            shutil.copy2(source, input_dir / f"sample_{index:02d}.txt")

        batch_workbook = self.copy_file(TXT_TEMPLATE, "batch.xlsx")
        result = run_job(
            JobConfig(
                data_kind=DATA_KIND_TXT,
                source_kind=SOURCE_KIND_DIRECTORY,
                input_path=str(input_dir),
                excel_path=str(batch_workbook),
                template_mode=True,
            )
        )

        self.assertEqual(result.processed_files, 9)
        self.assertEqual(len(result.output_paths), 2)
        for output_path in result.output_paths:
            self.assertTrue(Path(output_path).exists())

    def test_run_job_s2p_single_template_creates_output_workbook(self):
        s2p_file = self.copy_file(S2P_SAMPLE)
        template_file = self.copy_file(S2P_TEMPLATE, "s2p-template.xlsx")
        output_file = self.temp_dir / "s2p-result.xlsx"

        result = run_job(
            JobConfig(
                data_kind=DATA_KIND_S2P,
                source_kind=SOURCE_KIND_FILE,
                input_path=str(s2p_file),
                excel_path=str(template_file),
                template_mode=True,
                output_path=str(output_file),
            )
        )

        self.assertEqual(result.processed_files, 1)
        self.assertTrue(output_file.exists())

        workbook = openpyxl.load_workbook(output_file)
        self.assertIn(DEFAULT_TEMPLATE_SHEET, workbook.sheetnames)

    def test_run_job_s2p_directory_template_creates_output_workbooks(self):
        input_dir = self.temp_dir / "s2p-batch"
        input_dir.mkdir()
        shutil.copy2(S2P_SAMPLE, input_dir / "sample_01.s2p")
        shutil.copy2(S2P_SAMPLE, input_dir / "sample_02.s2p")
        template_file = self.copy_file(S2P_TEMPLATE, "s2p-batch-template.xlsx")

        result = run_job(
            JobConfig(
                data_kind=DATA_KIND_S2P,
                source_kind=SOURCE_KIND_DIRECTORY,
                input_path=str(input_dir),
                excel_path=str(template_file),
                template_mode=True,
            )
        )

        self.assertEqual(result.processed_files, 2)
        self.assertEqual(len(result.output_paths), 2)
        for output_path in result.output_paths:
            self.assertTrue(Path(output_path).exists())


if __name__ == "__main__":
    unittest.main()
