"""
AutoGraph Service - Сервис для переноса данных из TXT/S2P файлов в Excel.

Поддерживает режимы работы:
1. Обычный режим: создание нового листа с данными
2. Режим шаблона: создание нового Excel файла с копированием листа шаблона

Поддерживаемые форматы файлов:
- TXT: файлы с измерительными данными
- S2P: файлы S-параметров (Touchstone формат)

Использование:
    python main.py --txt <путь_к_TXT> --excel <путь_к_Excel> [--sheet <имя_листа>] [--output <путь_вывода>]
    python main.py --txt <путь_к_TXT> --excel <путь_к_Excel> --template [--template-sheet <имя_шаблона>] [--charts-sheet <имя_графиков>]
    python main.py --s2p <путь_к_S2P> --excel <путь_к_Excel> --template [--template-sheet <имя_шаблона>] [--output <путь_вывода>]
"""

import argparse
import sys
import os
import re
import shutil
from datetime import datetime
from typing import List, Dict

from autograph_service.txt_parser import TxtFileParser
from autograph_service.s2p_parser import S2PParser
from autograph_service.excel_handler import ExcelHandler
from autograph_service.s2p_excel_handler import S2PExcelHandler


def parse_args():
    """Парсит аргументы командной строки."""
    parser = argparse.ArgumentParser(
        description='Перенос данных из TXT/S2P файлов измерений в Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры использования:
  # Обычный режим - создание нового листа с данными
  python main.py --txt data.txt --excel report.xlsx
  
  # Режим шаблона - копирование шаблона и вставка данных
  python main.py --txt data.txt --excel template.xlsx --template
  
  # S2P режим шаблона - создание нового файла с данными (автоматическое имя)
  python main.py --s2p data.s2p --excel template.xlsx --template
  
  # S2P режим шаблона - с явным указанием выходного файла
  python main.py --s2p data.s2p --excel template.xlsx --template --output result.xlsx
  
  # Пакетная обработка S2P файлов
  python main.py --s2p-dir ./s2p_files --excel template.xlsx --template
  
  # С указанием имени шаблона
  python main.py --txt data.txt --excel template.xlsx --template --template-sheet "№1"
  
  # С указанием имени листа для графиков
  python main.py --txt data.txt --excel template.xlsx --template --charts-sheet "Графики"
  
  # Без интерактивных вопросов (автоматический режим)
  python main.py --s2p data.s2p --excel template.xlsx --template --no-prompt
        """
    )
    
    parser.add_argument(
        '--txt',
        default=None,
        help='Путь к TXT файлу с данными измерений'
    )
    
    parser.add_argument(
        '--s2p',
        default=None,
        help='Путь к S2P файлу с S-параметрами (Touchstone формат)'
    )
    
    parser.add_argument(
        '--txt-dir',
        default=None,
        dest='txt_dir',
        help='Путь к папке с TXT файлами для пакетной обработки'
    )
    
    parser.add_argument(
        '--s2p-dir',
        default=None,
        dest='s2p_dir',
        help='Путь к папке с S2P файлами для пакетной обработки'
    )
    
    parser.add_argument(
        '--excel',
        required=True,
        help='Путь к Excel файлу для записи данных'
    )
    
    parser.add_argument(
        '--sheet',
        default=None,
        help='Имя листа в Excel (по умолчанию: имя TXT файла без расширения)'
    )
    
    parser.add_argument(
        '--output',
        default=None,
        help='Путь к выходному Excel файлу (по умолчанию: имя S2P файла с расширением .xlsx)'
    )
    
    parser.add_argument(
        '--template',
        action='store_true',
        help='Режим шаблона: копирует шаблон листа и вставляет данные'
    )
    
    parser.add_argument(
        '--template-sheet',
        default='№1',
        dest='template_sheet',
        help='Имя листа-шаблона для копирования (по умолчанию: "№1")'
    )
    
    parser.add_argument(
        '--charts-sheet',
        default='Графики',
        dest='charts_sheet',
        help='Имя листа с графиками (по умолчанию: "Графики")'
    )
    
    parser.add_argument(
        '--create-chart',
        action='append',
        nargs=3,
        metavar=('Y_COLUMN', 'COLUMN_LETTER', 'Y_AXIS_TITLE'),
        help='Создать новый график. Формат: --create-chart <Y_колонка> <Буква_колонки> <Заголовок_Y_оси>. '
             'Y_колонка: Pout_W, Pout_dBm, Id_A, Ig_mA, Gain_dB, Efficiency. '
             'Пример: --create-chart Pout_W C "Pвых, Вт"'
    )
    
    parser.add_argument(
        '--no-prompt',
        action='store_true',
        help='Отключить интерактивные вопросы (использовать значения по умолчанию)'
    )
    
    parser.add_argument(
        '--charts-sheet-template',
        default='Графики',
        dest='charts_sheet_template',
        help='Имя листа с графиками в шаблоне (по умолчанию: "Графики")'
    )
    
    args = parser.parse_args()
    
    # Валидация: нужен либо --txt, либо --txt-dir, либо --s2p, либо --s2p-dir
    if not args.txt and not args.txt_dir and not args.s2p and not args.s2p_dir:
        parser.error("Требуется указать либо --txt/--s2p (один файл), либо --txt-dir/--s2p-dir (папка с файлами)")
    
    return args


def scan_txt_files(directory: str) -> List[str]:
    """
    Сканирует папку и возвращает список всех TXT файлов, отсортированных по имени.
    
    Args:
        directory: Путь к папке с TXT файлами
        
    Returns:
        Отсортированный список путей к TXT файлам
    """
    if not os.path.isdir(directory):
        raise NotADirectoryError(f"Указанный путь не является папкой: {directory}")
    
    txt_files = []
    for filename in os.listdir(directory):
        if filename.lower().endswith('.txt'):
            txt_files.append(os.path.join(directory, filename))
    
    # Сортируем по имени файла для предсказуемого порядка
    txt_files.sort(key=lambda x: os.path.basename(x))
    return txt_files


def scan_s2p_files(directory: str) -> List[str]:
    """
    Сканирует папку и возвращает список всех S2P файлов, отсортированных по имени.
    
    Args:
        directory: Путь к папке с S2P файлами
        
    Returns:
        Отсортированный список путей к S2P файлам
    """
    if not os.path.isdir(directory):
        raise NotADirectoryError(f"Указанный путь не является папкой: {directory}")
    
    s2p_files = []
    for filename in os.listdir(directory):
        if filename.lower().endswith('.s2p'):
            s2p_files.append(os.path.join(directory, filename))
    
    # Сортируем по имени файла для предсказуемого порядка
    s2p_files.sort(key=lambda x: os.path.basename(x))
    return s2p_files


def group_files(txt_files: List[str], max_per_group: int = 8) -> List[List[str]]:
    """
    Разбивает список TXT файлов на группы по max_per_group файлов.
    
    Args:
        txt_files: Список путей к TXT файлам
        max_per_group: Максимальное количество файлов в группе (по умолчанию 8)
        
    Returns:
        Список групп файлов
    """
    groups = []
    for i in range(0, len(txt_files), max_per_group):
        groups.append(txt_files[i:i + max_per_group])
    return groups


def generate_output_filename(original_excel_path: str, group_index: int, total_groups: int) -> str:
    """
    Генерирует имя выходного Excel файла для группы.
    
    Args:
        original_excel_path: Путь к исходному Excel файлу
        group_index: Индекс группы (0-based)
        total_groups: Общее количество групп
        
    Returns:
        Путь к выходному Excel файлу
    """
    directory = os.path.dirname(original_excel_path) or '.'
    basename = os.path.basename(original_excel_path)
    name, ext = os.path.splitext(basename)
    
    if total_groups == 1:
        # Если группа одна, используем оригинальное имя
        return original_excel_path
    
    # Создаем имя вида: original_name_part1.xlsx, original_name_part2.xlsx
    return os.path.join(directory, f"{name}_part{group_index + 1}{ext}")


def generate_sheet_name(file_path: str) -> str:
    """
    Генерирует имя листа на основе имени файла.
    
    Args:
        file_path: Путь к файлу (TXT или S2P)
        
    Returns:
        Имя листа (максимум 31 символ, без запрещенных символов)
    """
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    # Excel ограничивает имя листа 31 символом и запрещает: \ / ? * [ ] :
    invalid_chars = ['\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        base_name = base_name.replace(char, '_')
    # Обрезаем до 31 символа (Excel лимит)
    return base_name[:31]


def ask_s2p_template_questions(args, output_path: str) -> dict:
    """Задает интерактивные вопросы для одиночного S2P режима."""
    config = {}
    
    try:
        # Вопрос 1: Имя листа-шаблона
        config['template_sheet'] = input(f"Имя листа-шаблона [{args.template_sheet}]: ") or args.template_sheet
        
        # Вопрос 2: Копировать лист с графиками? (графики уже в шаблоне)
        create_charts = input("Графики уже есть в шаблоне, использовать их? [Да/нет]: ").lower()
        config['create_charts'] = create_charts != 'нет'
    except EOFError:
        # Если нет ввода (например, при пайпе), используем значения по умолчанию
        print("\nEOF detected, using defaults.")
        config['template_sheet'] = args.template_sheet
        config['create_charts'] = True
    
    config['output_path'] = output_path
    return config


def process_s2p_template(args) -> None:
    """
    Обрабатывает одиночный S2P файл в режиме шаблона.
    
    Новый принцип работы:
    1. Создается НОВЫЙ Excel файл (не копия всего шаблона)
    2. Копируется только указанный лист шаблона в новый файл
    3. Записываются S2P данные в скопированный лист
    4. Графики уже есть в листе шаблона - настраивается масштаб осей
    """
    # Определяем путь к выходному файлу
    output_path = generate_output_excel_name(args.s2p, args.output)
    
    # Задаем интерактивные вопросы или используем значения по умолчанию
    if args.no_prompt:
        config = {
            'output_path': output_path,
            'template_sheet': args.template_sheet,
            'create_charts': True,
        }
        print(f"Автоматический режим (--no-prompt):")
        print(f"  Выходной файл: {config['output_path']}")
        print(f"  Шаблон: {config['template_sheet']}")
    else:
        config = ask_s2p_template_questions(args, output_path)
    
    # Парсим S2P файл
    print(f"\nПарсинг S2P файла: {args.s2p}")
    parser = S2PParser()
    data = parser.parse_file(args.s2p)
    
    print(f"  Найдено строк данных: {len(data.rows)}")
    print(f"  Частота: {data.rows[0][0]} - {data.rows[-1][0]} GHz")
    
    # Загружаем исходный Excel шаблон
    template_handler = S2PExcelHandler(args.excel)
    template_handler.load_workbook()
    
    # Проверяем существование шаблона
    if config['template_sheet'] not in template_handler.get_sheet_names():
        print(f"\nОшибка: Шаблон листа '{config['template_sheet']}' не найден")
        print(f"  Доступные листы: {template_handler.get_sheet_names()}")
        sys.exit(1)
    
    # Создаем НОВЫЙ Excel файл
    print(f"\nСоздание нового Excel файла: {config['output_path']}")
    import openpyxl
    new_workbook = openpyxl.Workbook()
    
    # Удаляем стандартный лист, созданный по умолчанию
    if 'Sheet' in new_workbook.sheetnames:
        del new_workbook['Sheet']
    
    # Копируем лист шаблона из исходного файла в новый
    sheet_name = config['template_sheet'][:31]  # Обрезаем до 31 символа
    print(f"Копирование листа '{config['template_sheet']}' -> '{sheet_name}'")
    new_ws = template_handler.copy_template_sheet_to_new_workbook(
        source_sheet_name=config['template_sheet'],
        new_workbook=new_workbook,
        new_sheet_name=sheet_name
    )
    
    # Создаем обработчик для нового файла
    handler = S2PExcelHandler(config['output_path'])
    handler.workbook = new_workbook
    
    # Записываем S2P данные (начиная со строки 2, как в шаблоне)
    rows_written = handler.write_s2p_data(sheet_name, data.rows, start_data_row=2)
    print(f"  Записано S2P данных: {rows_written} строк")
    
    # Анализируем данные и масштабируем графики
    # Графики уже есть в листе шаблона, нужно только настроить масштаб
    if config['create_charts']:
        analysis = handler.analyze_s2p_data(data.rows)
        
        # Обновляем диапазоны данных в графиках (начиная со строки 2)
        handler.update_chart_data_ranges(sheet_name, sheet_name, len(data.rows), start_data_row=2, s2p_mode=True)
        print(f"  Диапазоны данных в графиках обновлены")
        
        # Автомасштабируем графики
        handler.auto_scale_charts(sheet_name, analysis)
        print(f"  Графики автоматически масштабированы")
        print(f"  Частота: {analysis['freq_min']:.4f} - {analysis['freq_max']:.4f} GHz")
        print(f"  S11: {analysis['s11_min']:.2f} - {analysis['s11_max']:.2f} dB")
        print(f"  S21: {analysis['s21_min']:.2f} - {analysis['s21_max']:.2f} dB")
    
    # Сохраняем новый файл
    handler.save(config['output_path'])
    print(f"\nСохранено в: {config['output_path']}")
    print(f"Листы в файле: {handler.get_sheet_names()}")


def process_s2p_batch(args) -> None:
    """
    Пакетная обработка папки с S2P файлами.
    
    Новый принцип работы:
    1. Для каждого S2P файла создается НОВЫЙ Excel файл
    2. Копируется только указанный лист шаблона
    3. Графики уже есть в листе шаблона - настраивается масштаб
    """
    # Сканируем папку
    s2p_files = scan_s2p_files(args.s2p_dir)
    
    if not s2p_files:
        print(f"Ошибка: В папке '{args.s2p_dir}' не найдено S2P файлов")
        sys.exit(1)
    
    print(f"Найдено S2P файлов: {len(s2p_files)}")
    
    # Задаем интерактивные вопросы или используем значения по умолчанию
    if args.no_prompt:
        config = {
            'template_sheet': args.template_sheet,
            'create_charts': True,
        }
        print(f"\nАвтоматический режим (--no-prompt):")
        print(f"  Шаблон: {config['template_sheet']}")
    else:
        config = ask_s2p_batch_questions(args)
    
    # Загружаем исходный Excel шаблон один раз
    template_handler = S2PExcelHandler(args.excel)
    template_handler.load_workbook()
    
    # Проверяем существование шаблона
    if config['template_sheet'] not in template_handler.get_sheet_names():
        print(f"\nОшибка: Шаблон листа '{config['template_sheet']}' не найден")
        print(f"  Доступные листы: {template_handler.get_sheet_names()}")
        sys.exit(1)
    
    print(f"\n{'=' * 50}")
    
    # Обрабатываем каждый S2P файл отдельно
    for s2p_idx, s2p_file in enumerate(s2p_files):
        print(f"\n{'=' * 50}")
        print(f"Файл {s2p_idx + 1}/{len(s2p_files)}")
        
        # Определяем путь к выходному файлу
        output_path = generate_output_excel_name(s2p_file, args.output)
        # Если пакетная обработка, добавляем индекс к имени
        if len(s2p_files) > 1:
            base, ext = os.path.splitext(output_path)
            output_path = f"{base}_{s2p_idx + 1}{ext}"
        
        print(f"\n  Обработка: {s2p_file}")
        print(f"  Выходной файл: {output_path}")
        
        # Парсим S2P
        parser = S2PParser()
        data = parser.parse_file(s2p_file)
        
        # Создаем НОВЫЙ Excel файл
        import openpyxl
        new_workbook = openpyxl.Workbook()
        
        # Удаляем стандартный лист, созданный по умолчанию
        if 'Sheet' in new_workbook.sheetnames:
            del new_workbook['Sheet']
        
        # Генерируем имя листа
        sheet_name = generate_unique_sheet_name([], os.path.splitext(os.path.basename(s2p_file))[0])
        
        # Копируем лист шаблона из исходного файла в новый
        print(f"  Копирование листа '{config['template_sheet']}' -> '{sheet_name}'")
        new_ws = template_handler.copy_template_sheet_to_new_workbook(
            source_sheet_name=config['template_sheet'],
            new_workbook=new_workbook,
            new_sheet_name=sheet_name
        )
        
        # Создаем обработчик для нового файла
        handler = S2PExcelHandler(output_path)
        handler.workbook = new_workbook
        
        # Записываем S2P данные (начиная со строки 2, как в шаблоне)
        rows_written = handler.write_s2p_data(sheet_name, data.rows, start_data_row=2)
        
        # Анализируем данные
        analysis = handler.analyze_s2p_data(data.rows)
        
        print(f"    Лист: {sheet_name}, строк: {rows_written}")
        
        # Масштабируем графики (если нужно)
        if config['create_charts']:
            # Обновляем диапазоны данных в графиках (начиная со строки 2)
            handler.update_chart_data_ranges(sheet_name, sheet_name, len(data.rows), start_data_row=2, s2p_mode=True)
            
            # Автомасштабируем графики
            handler.auto_scale_charts(sheet_name, analysis)
            print(f"    Графики автоматически масштабированы")
        
        # Сохраняем
        handler.save(output_path)
        print(f"  Сохранено в: {output_path}")
        print(f"  Листы: {handler.get_sheet_names()}")
    
    print(f"\n{'=' * 50}")
    print(f"Пакетная обработка завершена!")
    print(f"Создано файлов: {len(s2p_files)}")


def ask_s2p_batch_questions(args) -> dict:
    """Задает интерактивные вопросы для пакетной обработки S2P."""
    config = {}
    
    try:
        # Вопрос 1: Имя листа-шаблона
        config['template_sheet'] = input(f"Имя листа-шаблона [{args.template_sheet}]: ") or args.template_sheet
        
        # Вопрос 2: Использовать графики из шаблона?
        create_charts = input("Графики уже есть в шаблоне, использовать их? [Да/нет]: ").lower()
        config['create_charts'] = create_charts != 'нет'
    except EOFError:
        # Если нет ввода, используем значения по умолчанию
        print("\nEOF detected, using defaults.")
        config['template_sheet'] = args.template_sheet
        config['create_charts'] = True
    
    return config


def combine_s2p_analysis(analysis_dict: Dict[str, Dict]) -> Dict:
    """Объединяет анализ нескольких S2P файлов для масштабирования."""
    if not analysis_dict:
        return {
            'freq_min': 0, 'freq_max': 1,
            's11_min': -100, 's11_max': 0,
            's21_min': -100, 's21_max': 0,
            's12_min': -100, 's12_max': 0,
            's22_min': -100, 's22_max': 0,
            'num_rows': 0
        }
    
    combined = {
        'freq_min': float('inf'),
        'freq_max': float('-inf'),
        's11_min': float('inf'),
        's11_max': float('-inf'),
        's21_min': float('inf'),
        's21_max': float('-inf'),
        's12_min': float('inf'),
        's12_max': float('-inf'),
        's22_min': float('inf'),
        's22_max': float('-inf'),
        'num_rows': 0
    }
    
    for analysis in analysis_dict.values():
        combined['freq_min'] = min(combined['freq_min'], analysis['freq_min'])
        combined['freq_max'] = max(combined['freq_max'], analysis['freq_max'])
        combined['s11_min'] = min(combined['s11_min'], analysis['s11_min'])
        combined['s11_max'] = max(combined['s11_max'], analysis['s11_max'])
        combined['s21_min'] = min(combined['s21_min'], analysis['s21_min'])
        combined['s21_max'] = max(combined['s21_max'], analysis['s21_max'])
        combined['s12_min'] = min(combined['s12_min'], analysis['s12_min'])
        combined['s12_max'] = max(combined['s12_max'], analysis['s12_max'])
        combined['s22_min'] = min(combined['s22_min'], analysis['s22_min'])
        combined['s22_max'] = max(combined['s22_max'], analysis['s22_max'])
        combined['num_rows'] += analysis['num_rows']
    
    return combined


def generate_s2p_output_path(excel_path: str, group_idx: int, total_groups: int) -> str:
    """Генерирует имя выходного файла для группы S2P."""
    directory = os.path.dirname(excel_path) or '.'
    basename = os.path.basename(excel_path)
    name, ext = os.path.splitext(basename)
    
    if total_groups == 1:
        return os.path.join(directory, f"{name}_s2p{ext}")
    else:
        return os.path.join(directory, f"{name}_s2p_part{group_idx + 1}{ext}")


def generate_output_excel_name(s2p_path: str, output_arg: str = None) -> str:
    """
    Генерирует имя выходного Excel файла на основе S2P файла.
    
    Args:
        s2p_path: Путь к S2P файлу
        output_arg: Явно указанное имя выходного файла (--output)
        
    Returns:
        Путь к выходному Excel файлу
    """
    if output_arg:
        # Если пользователь явно указал --output, используем его
        result = output_arg
    else:
        # Иначе используем имя S2P файла с расширением .xlsx
        base_name = os.path.splitext(os.path.basename(s2p_path))[0]
        directory = os.path.dirname(s2p_path) or '.'
        result = os.path.join(directory, f"{base_name}.xlsx")
    
    # Добавляем расширение .xlsx если его нет
    if not result.endswith('.xlsx'):
        result += '.xlsx'
    
    return result


def process_single_file(args, file_path: str, excel_handler: ExcelHandler,
                        existing_sheets: list, is_first_file: bool = True) -> bool:
    """
    Обрабатывает один файл (TXT или S2P) и записывает данные в Excel.
    
    Args:
        args: Аргументы командной строки
        file_path: Путь к файлу (TXT или S2P)
        excel_handler: Обработчик Excel
        existing_sheets: Список существующих листов (до обработки)
        is_first_file: Является ли первым файлом (для определения output_path)
        
    Returns:
        True если обработка успешна
    """
    sheet_name = args.sheet or generate_sheet_name(file_path)
    
    # Определяем тип файла
    is_s2p = file_path.lower().endswith('.s2p')
    file_type = "S2P" if is_s2p else "TXT"
    
    print(f"\n  Обработка файла: {file_path}")
    print(f"  Тип файла: {file_type}")
    print(f"  Целевой лист: {sheet_name}")
    
    # Парсим файл в зависимости от типа
    if is_s2p:
        parser = S2PParser()
        measurement_data = parser.parse_file(file_path)
        headers = measurement_data.headers
        rows = measurement_data.rows
        metadata = measurement_data.metadata
        
        # Для S2P используем свои заголовки
        headers = parser.get_headers_for_template()
        
        if not rows:
            print("  Предупреждение: Данные не найдены в S2P файле")
    else:
        parser = TxtFileParser()
        measurement_data = parser.parse_file(file_path)
        headers = measurement_data.headers
        rows = measurement_data.rows
        metadata = measurement_data.metadata
        
        if not headers:
            print("  Предупреждение: Заголовки таблицы не найдены в TXT файле")
    
    print(f"  Найдено строк данных: {len(rows)}")
    print(f"  Количество колонок: {len(headers)}")
    if metadata:
        print(f"  Метаданные: {list(metadata.keys())}")
    
    # Инициализируем new_sheet_name заранее для использования в выводе
    new_sheet_name = None
    
    if args.template:
        # РЕЖИМ ШАБЛОНА: копируем шаблон и вставляем данные
        
        # Проверяем существование шаблона
        if args.template_sheet not in excel_handler.get_sheet_names():
            print(f"  Ошибка: Шаблон листа '{args.template_sheet}' не найден")
            print(f"  Доступные листы: {excel_handler.get_sheet_names()}")
            return False
        
        # Создаем новый лист из шаблона (для всех типов файлов)
        new_sheet_name = generate_unique_sheet_name(excel_handler.get_sheet_names(), sheet_name)
        print(f"  Новый лист: {new_sheet_name}")
        
        # Копируем шаблон
        excel_handler.copy_template_sheet(
            source_sheet_name=args.template_sheet,
            new_sheet_name=new_sheet_name
        )
        print(f"  Шаблон '{args.template_sheet}' скопирован в '{new_sheet_name}'")
        
        # Записываем данные в новый лист в зависимости от типа файла
        if is_s2p:
            # Для S2P записываем данные в колонки A-E (Freq, S11M, S21M, S12M, S22M)
            rows_written = excel_handler.write_s2p_data_to_template_sheet(
                sheet_name=new_sheet_name,
                data_rows=rows,
                start_data_row=3  # S2P данные начинаются со строки 3
            )
            print(f"  Записано S2P данных: {rows_written} строк")
        else:
            # Для TXT записываем данные в колонки A-H (основная таблица)
            rows_written = excel_handler.write_data_to_template_sheet(
                sheet_name=new_sheet_name,
                data_rows=rows,
                start_data_row=11
            )
            print(f"  Записано в A-H: {rows_written} строк")
            
            # Записываем данные в колонки P-U (значения для графиков)
            excel_handler.write_data_columns_pu(
                sheet_name=new_sheet_name,
                data_rows=rows,
                start_row=11
            )
            print(f"  Данные записаны в P-U: {rows_written} строк")
            
            # Записываем метаданные в строку 1 (для формул шаблона)
            excel_handler.write_metadata_to_template_sheet(
                sheet_name=new_sheet_name,
                metadata=metadata
            )
            print(f"  Метаданные записаны в строку 1")
        
        # Добавляем серию на лист Графики (только для TXT файлов)
        if args.charts_sheet in excel_handler.get_sheet_names():
            if not is_s2p:
                result = excel_handler.add_chart_series(
                    charts_sheet_name=args.charts_sheet,
                    new_data_sheet_name=new_sheet_name,
                    series_name=new_sheet_name,
                    num_data_rows=rows_written
                )
                if result is not None:
                    print(f"  Серия '{new_sheet_name}' добавлена на лист '{args.charts_sheet}' (слот {result + 1})")
                
                # Создаем новые графики если указано (только для первого файла)
                if args.create_chart and is_first_file:
                    print(f"  Создание новых графиков...")
                    # Маппинг позиций для новых графиков
                    chart_positions = [
                        (1, 92),   # A92 - первый новый график
                        (7, 92),   # G92
                        (13, 92),  # M92
                        (19, 92),  # S92
                        (25, 92),  # Y92
                        (31, 92),  # AE92
                        (37, 92),  # AK92
                        (43, 92),  # AQ92
                    ]
                    
                    for i, (y_col, col_letter, y_title) in enumerate(args.create_chart):
                        if i < len(chart_positions):
                            pos_col, pos_row = chart_positions[i]
                            excel_handler.create_new_chart(
                                charts_sheet_name=args.charts_sheet,
                                chart_title=y_title,
                                y_column_name=y_col,
                                y_column_letter=col_letter,
                                y_axis_title=y_title,
                                position_col=pos_col,
                                position_row=pos_row,
                            )
                            print(f"  Создан график '{y_title}' (колонка {col_letter})")
            else:
                print(f"  Для S2P файлов добавление серий на лист графиков не поддерживается")
        else:
            print(f"  Предупреждение: Лист '{args.charts_sheet}' не найден, графики не обновлены")
    else:
        # ОБЫЧНЫЙ РЕЖИМ: создаем новый лист с данными
        
        if sheet_name in excel_handler.get_sheet_names():
            print(f"  Лист '{sheet_name}' будет перезаписан")
        
        # Записываем данные
        excel_handler.overwrite_sheet(
            sheet_name=sheet_name,
            headers=headers,
            rows=rows
        )
    
    return True


def main():
    """Основная функция сервиса."""
    args = parse_args()
    
    # Определяем тип файлов для обработки
    is_s2p_mode = args.s2p or args.s2p_dir
    file_type = "S2P" if is_s2p_mode else "TXT"
    
    print(f"AutoGraph Service - Перенос данных из {file_type} в Excel")
    print(f"{'=' * 50}")
    print(f"Excel шаблон: {args.excel}")
    print(f"Режим шаблона: {'Да' if args.template else 'Нет'}")
    if args.template:
        print(f"Шаблон листа: {args.template_sheet}")
        if is_s2p_mode:
            # В новом режиме графики уже есть в шаблоне
            print(f"  (графики уже есть в листе шаблона)")
        else:
            print(f"Лист графиков: {args.charts_sheet}")
    if args.create_chart:
        print(f"Создание графиков: {args.create_chart}")
    if args.output:
        print(f"Выходной файл: {args.output}")
    
    try:
        # S2P режим шаблона - отдельная обработка
        if is_s2p_mode and args.template:
            if args.s2p_dir:
                # Пакетная обработка S2P файлов
                process_s2p_batch(args)
            elif args.s2p:
                # Одиночный S2P файл
                process_s2p_template(args)
            return
        
        # Определяем режим работы: один файл или папка
        if args.txt_dir or args.s2p_dir:
            # ПАКЕТНЫЙ РЕЖИМ: обработка всех файлов в папке (TXT)
            directory = args.txt_dir
            print(f"Режим: Пакетная обработка папки")
            print(f"Папка с TXT файлами: {directory}")
            
            # Сканируем папку
            files = scan_txt_files(directory)
            
            if not files:
                print(f"Ошибка: В папке '{directory}' не найдено TXT файлов")
                sys.exit(1)
            
            print(f"Найдено TXT файлов: {len(files)}")
            
            # Группируем файлы по 8 штук
            file_groups = group_files(files, max_per_group=8)
            total_groups = len(file_groups)
            print(f"Групп файлов (по 8 максимум): {total_groups}")
            print(f"{'=' * 50}")
            
            # Обрабатываем каждую группу
            for group_idx, current_group in enumerate(file_groups):
                print(f"\n{'=' * 50}")
                print(f"Группа {group_idx + 1}/{total_groups} ({len(current_group)} файлов)")
                
                # Определяем путь к выходному файлу
                output_path = generate_output_filename(args.excel, group_idx, total_groups)
                
                # Для первой группы используем оригинальный Excel, для остальных - копию
                if group_idx == 0:
                    current_excel_path = args.excel
                else:
                    prev_output = generate_output_filename(args.excel, group_idx - 1, total_groups)
                    shutil.copy2(prev_output, output_path)
                    current_excel_path = output_path
                    print(f"Создана копия Excel файла: {output_path}")
                
                print(f"Excel файл: {current_excel_path}")
                print(f"Сохранение в: {output_path}")
                
                # Загружаем Excel
                excel_handler = ExcelHandler(current_excel_path)
                excel_handler.load_workbook()
                
                # Обрабатываем каждый файл в группе
                for file_idx, file_path in enumerate(current_group):
                    is_first_file = (group_idx == 0 and file_idx == 0)
                    success = process_single_file(
                        args, file_path, excel_handler,
                        excel_handler.get_sheet_names(),
                        is_first_file
                    )
                    if not success:
                        print(f"Ошибка при обработке файла: {file_path}")
                        sys.exit(1)
                
                # Сохраняем результат группы
                excel_handler.save(output_path)
                print(f"Сохранено в: {output_path}")
                print(f"Листы в файле: {excel_handler.get_sheet_names()}")
            
            print(f"\n{'=' * 50}")
            print(f"Пакетная обработка завершена!")
            print(f"Создано файлов: {total_groups}")
            for i in range(total_groups):
                path = generate_output_filename(args.excel, i, total_groups)
                print(f"  - {path}")
            print("Готово!")
            
        elif args.txt or args.s2p:
            # ОДИНОЧНЫЙ РЕЖИМ: обработка одного файла
            file_path = args.txt if args.txt else args.s2p
            if not os.path.exists(file_path):
                print(f"Ошибка: {file_type} файл не найден: {file_path}")
                sys.exit(1)
            
            sheet_name = args.sheet or generate_sheet_name(file_path)
            print(f"Режим: Одиночный файл")
            print(f"{file_type} файл: {file_path}")
            if not args.template:
                print(f"Имя листа: {sheet_name}")
            print(f"{'=' * 50}")
            
            # Загружаем Excel
            excel_handler = ExcelHandler(args.excel)
            excel_handler.load_workbook()
            
            # Обрабатываем файл
            success = process_single_file(
                args, file_path, excel_handler,
                excel_handler.get_sheet_names(),
                is_first_file=True
            )
            
            if not success:
                print(f"Ошибка при обработке файла: {file_path}")
                sys.exit(1)
            
            # Сохраняем
            output_path = args.output if args.output else None
            excel_handler.save(output_path)
            
            if output_path:
                print(f"\nДанные сохранены в: {output_path}")
            else:
                print(f"\nДанные сохранены в: {args.excel}")
            print("\nГотово!")
        
    except FileNotFoundError as e:
        print(f"\nОшибка: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\nНеожиданная ошибка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def generate_unique_sheet_name(existing_sheets: list, base_name: str) -> str:
    """
    Генерирует уникальное имя листа, добавляя суффикс если нужно.
    
    Args:
        existing_sheets: Список существующих имен листов
        base_name: Базовое имя листа
        
    Returns:
        Уникальное имя листа (максимум 31 символ)
    """
    # Обрезаем базовое имя до 31 символа
    base_name = base_name[:31]
    
    if base_name not in existing_sheets:
        return base_name
    
    # Добавляем суффикс, учитывая лимит в 31 символ
    # Оставляем место для суффикса (_1, _2, и т.д.)
    max_base_len = 31 - 3  # "_1" = 2 символа + запас
    counter = 1
    while counter < 100:  # Защита от бесконечного цикла
        short_base = base_name[:max_base_len]
        new_name = f"{short_base}_{counter}"
        if new_name not in existing_sheets:
            return new_name[:31]
        counter += 1
    
    # Если не нашли уникальное имя, возвращаем что есть
    return base_name


if __name__ == '__main__':
    main()
