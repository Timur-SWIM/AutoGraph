"""
Модуль для обработки S2P шаблонов Excel.

Поддерживает:
- Копирование листов шаблона и графиков
- Запись S2P данных в шаблон
- Анализ S2P данных для масштабирования
- Обновление диапазонов данных в графиках
- Автоматическое масштабирование осей графиков
"""

import re
import copy
from typing import List, Optional, Dict, Tuple
import openpyxl
from openpyxl.utils import get_column_letter


class S2PExcelHandler:
    """Обработчик для S2P шаблонов Excel."""
    
    def __init__(self, file_path: str):
        """
        Инициализирует обработчик Excel файла.
        
        Args:
            file_path: Путь к Excel файлу
        """
        self.file_path = file_path
        self.workbook = None
    
    def load_workbook(self) -> None:
        """Загружает существующий Excel файл или создает новый."""
        import os
        if os.path.exists(self.file_path):
            self.workbook = openpyxl.load_workbook(self.file_path)
        else:
            self.workbook = openpyxl.Workbook()
    
    def get_sheet_names(self) -> List[str]:
        """Возвращает список имен листов в книге."""
        if self.workbook is None:
            self.load_workbook()
        return self.workbook.sheetnames
    
    def copy_template_sheet(
        self,
        source_sheet_name: str,
        new_sheet_name: str
    ) -> None:
        """
        Копирует лист-шаблон с сохранением формул и структуры.
        
        Args:
            source_sheet_name: Имя исходного листа-шаблона
            new_sheet_name: Имя нового листа
        """
        if self.workbook is None:
            self.load_workbook()
        
        if source_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Исходный лист '{source_sheet_name}' не найден")
        
        source_ws = self.workbook[source_sheet_name]
        
        # Обрезаем имя листа до 31 символа (лимит Excel)
        safe_sheet_name = new_sheet_name[:31]
        
        # Создаем новый лист
        new_ws = self.workbook.create_sheet(title=safe_sheet_name)
        
        # Копируем все ячейки вручную
        max_row = 501
        max_col = 21
        
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = source_ws.cell(row=row_idx, column=col_idx)
                coord = f"{get_column_letter(col_idx)}{row_idx}"
                new_cell = new_ws[coord]
                
                if cell.value is not None:
                    new_cell.value = cell.value
                
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy.copy(cell.alignment)
        
        # Копируем размеры колонок
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
            new_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        
        # Копируем размеры строк
        for row_num, row_dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height
            new_ws.row_dimensions[row_num].hidden = row_dim.hidden
        
        # Копируем объединенные ячейки
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))
    
    def copy_template_sheet_to_new_workbook(
        self,
        source_sheet_name: str,
        new_workbook,
        new_sheet_name: str
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        Копирует лист-шаблон из текущей книги в новую книгу.
        Сохраняет формулы, стили, графики и структуру.
        
        Args:
            source_sheet_name: Имя исходного листа-шаблона
            new_workbook: Новая книга openpyxl для копирования
            new_sheet_name: Имя нового листа
            
        Returns:
            Новый лист в новой книге
        """
        if self.workbook is None:
            self.load_workbook()
        
        if source_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Исходный лист '{source_sheet_name}' не найден")
        
        source_ws = self.workbook[source_sheet_name]
        
        # Обрезаем имя листа до 31 символа (лимит Excel)
        safe_sheet_name = new_sheet_name[:31]
        
        # Создаем новый лист в новой книге
        new_ws = new_workbook.create_sheet(title=safe_sheet_name)
        
        # Копируем все ячейки вручную
        max_row = 501
        max_col = 21
        
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = source_ws.cell(row=row_idx, column=col_idx)
                coord = f"{get_column_letter(col_idx)}{row_idx}"
                new_cell = new_ws[coord]
                
                if cell.value is not None:
                    new_cell.value = cell.value
                
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy.copy(cell.alignment)
        
        # Копируем размеры колонок
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
            new_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        
        # Копируем размеры строк
        for row_num, row_dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height
            new_ws.row_dimensions[row_num].hidden = row_dim.hidden
        
        # Копируем объединенные ячейки
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))
        
        # Копируем графики с привязками через внутренний API
        if hasattr(source_ws, '_charts') and source_ws._charts:
            from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
            from openpyxl.drawing.spreadsheet_drawing import AnchorClientData
            
            for chart in source_ws._charts:
                # Создаем глубокую копию графика
                new_chart = copy.deepcopy(chart)
                
                # Добавляем график на лист
                new_ws.add_chart(new_chart)
                
                # Находим привязку исходного графика и копируем её
                if hasattr(source_ws, 'drawing') and source_ws.drawing:
                    for anchor in source_ws.drawing.anchors:
                        if hasattr(anchor, 'graphicFrame') and anchor.graphicFrame:
                            # Проверяем, относится ли эта привязка к текущему графику
                            if hasattr(anchor, 'sp') and anchor.sp == chart:
                                # Создаем новую привязку для скопированного графика
                                new_anchor = TwoCellAnchor(
                                    _from=copy.deepcopy(anchor._from),
                                    to=copy.deepcopy(anchor.to),
                                    clientData=copy.deepcopy(anchor.clientData) if hasattr(anchor, 'clientData') else AnchorClientData()
                                )
                                new_anchor.sp = new_chart
                                # Заменяем автоматически созданную привязку
                                if new_ws.drawing:
                                    new_ws.drawing.anchors.append(new_anchor)
        
        return new_ws
    
    def copy_charts_sheet(
        self,
        source_sheet_name: str,
        new_sheet_name: str
    ) -> None:
        """
        Копирует лист с графиками из шаблона БЕЗ ИЗМЕНЕНИЙ структуры.
        Использует внутренний API openpyxl для правильного копирования графиков.
        
        Args:
            source_sheet_name: Имя исходного листа с графиками
            new_sheet_name: Имя нового листа с графиками
        """
        if self.workbook is None:
            self.load_workbook()
        
        if source_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Исходный лист графиков '{source_sheet_name}' не найден")
        
        source_ws = self.workbook[source_sheet_name]
        
        # Обрезаем имя листа до 31 символа
        safe_sheet_name = new_sheet_name[:31]
        
        # Создаем новый лист
        new_ws = self.workbook.create_sheet(title=safe_sheet_name)
        
        # Копируем все ячейки
        for row in source_ws.iter_rows():
            for cell in row:
                coord = cell.coordinate
                new_cell = new_ws[coord]
                
                if cell.value is not None:
                    new_cell.value = cell.value
                
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy.copy(cell.alignment)
        
        # Копируем размеры колонок
        for col_letter, col_dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = col_dim.width
            new_ws.column_dimensions[col_letter].hidden = col_dim.hidden
        
        # Копируем размеры строк
        for row_num, row_dim in source_ws.row_dimensions.items():
            new_ws.row_dimensions[row_num].height = row_dim.height
            new_ws.row_dimensions[row_num].hidden = row_dim.hidden
        
        # Копируем объединенные ячейки
        for merged_range in source_ws.merged_cells.ranges:
            new_ws.merge_cells(str(merged_range))
        
        # Копируем графики с привязками через внутренний API
        if hasattr(source_ws, '_charts') and source_ws._charts:
            from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
            from openpyxl.drawing.spreadsheet_drawing import AnchorClientData
            
            for chart in source_ws._charts:
                # Создаем глубокую копию графика
                new_chart = copy.deepcopy(chart)
                
                # Добавляем график на лист
                new_ws.add_chart(new_chart)
                
                # Находим привязку исходного графика и копируем её
                if hasattr(source_ws, 'drawing') and source_ws.drawing:
                    for anchor in source_ws.drawing.anchors:
                        if hasattr(anchor, 'graphicFrame') and anchor.graphicFrame:
                            # Проверяем, относится ли эта привязка к текущему графику
                            if hasattr(anchor, 'sp') and anchor.sp == chart:
                                # Создаем новую привязку для скопированного графика
                                new_anchor = TwoCellAnchor(
                                    _from=copy.deepcopy(anchor._from),
                                    to=copy.deepcopy(anchor.to),
                                    clientData=copy.deepcopy(anchor.clientData) if hasattr(anchor, 'clientData') else AnchorClientData()
                                )
                                new_anchor.sp = new_chart
                                # Заменяем автоматически созданную привязку
                                if new_ws.drawing:
                                    new_ws.drawing.anchors.append(new_anchor)
    
    def write_s2p_data(
        self,
        sheet_name: str,
        data_rows: List[List],
        start_data_row: int = 2
    ) -> int:
        """
        Записывает данные из S2P файла в лист-шаблон.
        
        Данные S2P записываются в столбцы A-E:
          A - Freq, GHz
          B - S11M (дБ)
          C - S21M (дБ)
          D - S12M (дБ)
          E - S22M (дБ)
        
        Args:
            sheet_name: Имя листа для записи
            data_rows: Строки данных из S2P файла
            start_data_row: Начальная строка для данных
            
        Returns:
            Количество записанных строк
        """
        if self.workbook is None:
            self.load_workbook()
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
        
        ws = self.workbook[sheet_name]
        
        rows_written = 0
        for i, row_data in enumerate(data_rows):
            target_row = start_data_row + i
            for col_idx, value in enumerate(row_data):
                if col_idx < 5:
                    cell_value = self._convert_value(value)
                    ws.cell(row=target_row, column=col_idx + 1, value=cell_value)
            rows_written += 1
        
        return rows_written
    
    def analyze_s2p_data(self, data_rows: List[List]) -> Dict:
        """
        Анализирует S2P данные и возвращает диапазоны для масштабирования.
        
        Args:
            data_rows: Строки данных из S2P файла
            
        Returns:
            Словарь с min/max значениями:
            {
                'freq_min': float, 'freq_max': float,
                's11_min': float, 's11_max': float,
                's21_min': float, 's21_max': float,
                's12_min': float, 's12_max': float,
                's22_min': float, 's22_max': float,
                'num_rows': int
            }
        """
        if not data_rows:
            return {
                'freq_min': 0, 'freq_max': 1,
                's11_min': -100, 's11_max': 0,
                's21_min': -100, 's21_max': 0,
                's12_min': -100, 's12_max': 0,
                's22_min': -100, 's22_max': 0,
                'num_rows': 0
            }
        
        freq_values = []
        s11_values = []
        s21_values = []
        s12_values = []
        s22_values = []
        
        for row in data_rows:
            if len(row) >= 5:
                try:
                    freq_values.append(float(row[0]))
                    s11_values.append(float(row[1]))
                    s21_values.append(float(row[2]))
                    s12_values.append(float(row[3]))
                    s22_values.append(float(row[4]))
                except (ValueError, TypeError):
                    pass
        
        return {
            'freq_min': min(freq_values) if freq_values else 0,
            'freq_max': max(freq_values) if freq_values else 1,
            's11_min': min(s11_values) if s11_values else -100,
            's11_max': max(s11_values) if s11_values else 0,
            's21_min': min(s21_values) if s21_values else -100,
            's21_max': max(s21_values) if s21_values else 0,
            's12_min': min(s12_values) if s12_values else -100,
            's12_max': max(s12_values) if s12_values else 0,
            's22_min': min(s22_values) if s22_values else -100,
            's22_max': max(s22_values) if s22_values else 0,
            'num_rows': len(data_rows)
        }
    
    def update_chart_data_ranges(
        self,
        charts_sheet_name: str,
        data_sheet_name: str,
        num_data_rows: int,
        start_data_row: int = 2,
        s2p_mode: bool = False
    ) -> None:
        """
        Обновляет диапазоны данных в формулах графиков.
        
        Находит все графики на листе и обновляет ссылки на данные:
        - X значения: Частота (колонка A листа данных)
        - Y значения: S11M, S21M, S12M, S22M (колонки B-E листа данных) для S2P
                     или оригинальные колонки для TXT
        
        Args:
            charts_sheet_name: Имя листа с графиками
            data_sheet_name: Имя листа с данными
            num_data_rows: Количество строк данных
            start_data_row: Начальная строка данных
            s2p_mode: Если True, обновляет ссылки на S2P колонки (B-E)
        """
        if self.workbook is None:
            self.load_workbook()
        
        if charts_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{charts_sheet_name}' не найден")
        
        charts_ws = self.workbook[charts_sheet_name]
        
        end_data_row = start_data_row + num_data_rows - 1
        
        # S2P колонки уже правильные: B=S11, C=S21, D=S12, E=S22
        # Не нужно маппить их, просто обновляем диапазон строк
        s2p_columns = {'B', 'C', 'D', 'E'}
        
        for chart in charts_ws._charts:
            for series_idx, series in enumerate(chart.series):
                # Обновляем X значения (Частота - колонка A)
                if hasattr(series, 'xVal') and series.xVal:
                    if hasattr(series.xVal, 'numRef') and series.xVal.numRef:
                        series.xVal.numRef.f = f"'{data_sheet_name}'!$A${start_data_row}:$A${end_data_row}"
                        # Очищаем кэшированные данные
                        self._clear_num_cache(series.xVal.numRef)
                
                # Обновляем Y значения
                if hasattr(series, 'yVal') and series.yVal:
                    if hasattr(series.yVal, 'numRef') and series.yVal.numRef:
                        current_ref = series.yVal.numRef.f
                        # Извлекаем букву колонки из текущей ссылки
                        match = re.search(r'\$([A-Z]+)\$', current_ref)
                        if match:
                            col_letter = match.group(1)
                            
                            # В S2P режиме: если колонка уже B, C, D или E - оставляем как есть
                            # Иначе маппим из TXT колонок
                            if s2p_mode:
                                if col_letter in s2p_columns:
                                    # Уже S2P колонка, не маппим
                                    target_col = col_letter
                                else:
                                    # TXT колонка, маппим на S2P
                                    txt_to_s2p_mapping = {
                                        'C': 'B',   # Pout_W -> S11
                                        'K': 'C',   # Pout_dBm -> S21
                                        'S': 'D',   # Id_A -> S12
                                        'AA': 'E',  # Ig_mA -> S22
                                        'AI': 'B',  # Gain_dB -> S11
                                        'AQ': 'C',  # Efficiency -> S21
                                    }
                                    target_col = txt_to_s2p_mapping.get(col_letter, col_letter)
                            else:
                                target_col = col_letter
                            
                            series.yVal.numRef.f = f"'{data_sheet_name}'!${target_col}${start_data_row}:${target_col}${end_data_row}"
                            # Очищаем кэшированные данные
                            self._clear_num_cache(series.yVal.numRef)
    
    def _clear_num_cache(self, num_ref) -> None:
        """Очищает кэшированные данные в ссылке на число."""
        if hasattr(num_ref, 'numCache') and num_ref.numCache:
            num_ref.numCache.ptCount = None
            num_ref.numCache.pt = []
    
    def _get_chart_title_text(self, chart) -> str:
        """Извлекает текстовый заголовок из объекта графика."""
        if not hasattr(chart, 'title') or not chart.title:
            return ''
        
        title_obj = chart.title
        # Пытаемся извлечь текст из различных структур
        if hasattr(title_obj, 'tx') and title_obj.tx:
            tx = title_obj.tx
            if hasattr(tx, 'strRef') and tx.strRef:
                if hasattr(tx.strRef, 'strCache') and tx.strRef.strCache:
                    if hasattr(tx.strRef.strCache, 'pt') and tx.strRef.strCache.pt:
                        pt = tx.strRef.strCache.pt[0]
                        if hasattr(pt, 'v'):
                            return str(pt.v)
            if hasattr(tx, 'rich') and tx.rich:
                if hasattr(tx.rich, 'p') and tx.rich.p:
                    texts = []
                    for p in tx.rich.p:
                        if hasattr(p, 'r') and p.r:
                            for r in p.r:
                                if hasattr(r, 't'):
                                    texts.append(r.t)
                    return ' '.join(texts)
        
        return ''
    
    def auto_scale_charts(
        self,
        charts_sheet_name: str,
        analysis: Dict
    ) -> None:
        """
        Автоматически масштабирует все графики на листе.
        
        Для каждого графика:
        - Определяет тип по заголовку (S11, S21, S12, S22)
        - Устанавливает X axis: freq_min..freq_max
        - Устанавливает Y axis: соответствующий s_param_min..s_param_max
        
        Args:
            charts_sheet_name: Имя листа с графиками
            analysis: Словарь с min/max значениями из analyze_s2p_data()
        """
        if self.workbook is None:
            self.load_workbook()
        
        if charts_sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{charts_sheet_name}' не найден")
        
        charts_ws = self.workbook[charts_sheet_name]
        
        # Маппинг заголовков графиков к типам данных
        chart_type_mapping = {
            'S11': ('s11_min', 's11_max'),
            'S21': ('s21_min', 's21_max'),
            'S12': ('s12_min', 's12_max'),
            'S22': ('s22_min', 's22_max'),
        }
        
        for chart_idx, chart in enumerate(charts_ws._charts):
            # Извлекаем текст заголовка графика
            chart_title_text = self._get_chart_title_text(chart)
            
            # Определяем тип графика по заголовку или по порядку
            chart_type = None
            for key, (min_key, max_key) in chart_type_mapping.items():
                if chart_title_text and key in chart_title_text.upper():
                    chart_type = (min_key, max_key)
                    break
            
            # Если не определили по заголовку, используем порядок графиков
            # Первые 4 графика = S11, S21, S12, S22
            # Графики 5-6 дублируют S11, S21
            if chart_type is None and chart_idx < 6:
                s2p_types = [('s11_min', 's11_max'), ('s21_min', 's21_max'),
                             ('s12_min', 's12_max'), ('s22_min', 's22_max'),
                             ('s11_min', 's11_max'), ('s21_min', 's21_max')]
                chart_type = s2p_types[chart_idx]
            
            if chart_type:
                from openpyxl.chart.axis import Scaling
                
                min_key, max_key = chart_type
                y_min = analysis.get(min_key, -100)
                y_max = analysis.get(max_key, 0)
                
                # Устанавливаем диапазон Y с запасом 10%
                y_range_min, y_range_max = self._calculate_axis_range(y_min, y_max)
                
                # Округляем для корректного отображения
                y_range_min = round(y_range_min, 2)
                y_range_max = round(y_range_max, 2)
                
                # Создаем или обновяем scaling для Y оси
                if not hasattr(chart.y_axis, 'scaling') or chart.y_axis.scaling is None:
                    chart.y_axis.scaling = Scaling()
                chart.y_axis.scaling.min = y_range_min
                chart.y_axis.scaling.max = y_range_max
                
                # Устанавливаем диапазон X (частота)
                freq_min = analysis.get('freq_min', 0)
                freq_max = analysis.get('freq_max', 1)
                freq_range_min, freq_range_max = self._calculate_axis_range(freq_min, freq_max)
                
                # Округляем для корректного отображения
                freq_range_min = round(freq_range_min, 4)
                freq_range_max = round(freq_range_max, 4)
                
                # Создаем или обновяем scaling для X оси
                if not hasattr(chart.x_axis, 'scaling') or chart.x_axis.scaling is None:
                    chart.x_axis.scaling = Scaling()
                chart.x_axis.scaling.min = freq_range_min
                chart.x_axis.scaling.max = freq_range_max
    
    def _calculate_axis_range(
        self,
        data_min: float,
        data_max: float,
        margin: float = 0.1
    ) -> Tuple[float, float]:
        """
        Вычисляет диапазон оси с запасом.
        
        Args:
            data_min: Минимальное значение данных
            data_max: Максимальное значение данных
            margin: Запас (10% по умолчанию)
            
        Returns:
            Кортеж (min_with_margin, max_with_margin)
        """
        range_val = data_max - data_min
        if range_val == 0:
            range_val = abs(data_min) if data_min != 0 else 1
            range_val = range_val * 0.1  # 10% от значения
        
        margin_val = range_val * margin
        return (data_min - margin_val, data_max + margin_val)
    
    def _convert_value(self, value):
        """Преобразует строковое значение в число где возможно."""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return value
        
        str_value = str(value).strip().replace(',', '.')
        
        try:
            if '.' not in str_value:
                return int(str_value)
        except (ValueError, TypeError):
            pass
        
        try:
            return float(str_value)
        except (ValueError, TypeError):
            return value
    
    def save(self, output_path: Optional[str] = None) -> None:
        """
        Сохраняет Excel файл.
        
        Args:
            output_path: Путь для сохранения (если None, сохраняет в исходный файл)
        """
        if self.workbook is None:
            raise ValueError("Книга не загружена. Вызовите load_workbook() сначала.")
        
        # Устанавливаем флаг автоматического пересчета формул
        self.workbook.calculation.calcOnSave = True
        self.workbook.calculation.calcId = 0
        self.workbook.calculation.fullCalcOnLoad = True
        
        save_path = output_path or self.file_path
        self.workbook.save(save_path)
